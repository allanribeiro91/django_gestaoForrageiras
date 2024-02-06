from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage
from django.contrib import auth, messages
from django.urls import reverse
from apps.main.models import CustomLog
from django.http import JsonResponse, HttpResponse
from apps.fornecedores.models import Fornecedores, Fornecedores_Comunicacoes, UF_Municipio, CNPJ_NATUREZA_JURIDICA, CNPJ_CNAE, Fornecedores_Faq, Fornecedores_Representantes
from apps.fornecedores.forms import FornecedoresForm, FornecedoresFaqForm, FornecedoresRepresentantesForm, FornecedoresComunicacoesForm
from apps.usuarios.models import Usuarios
from setup.choices import UNIDADE_DAF2, CNPJ_HIERARQUIA, CNPJ_PORTE, TIPO_DIREITO, FAQ_FORNECEDOR_TOPICO, CARGOS_FUNCOES, GENERO_SEXUAL, TIPO_COMUNICACAO, STATUS_ENVIO_COMUNICACAO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import json


#FORNECEDORES
def fornecedores(request):
    tab_fornecedores = Fornecedores.objects.filter(del_status=False).order_by('nome_fantasia')
    total_fornecedores = tab_fornecedores.count()
    lista_ufs = UF_Municipio.objects.values_list('uf_sigla', flat=True).distinct().order_by('uf_sigla')
    conteudo = {
         'tab_fornecedores': tab_fornecedores,
         'total_fornecedores': total_fornecedores,
         'lista_cnpj_hierarquia': CNPJ_HIERARQUIA,
         'lista_cnpj_porte': CNPJ_PORTE,
         'lista_tipo_direito': TIPO_DIREITO,
         'lista_ufs': lista_ufs,
    }
    return render(request, 'fornecedores/fornecedores.html', conteudo)

def fornecedor_ficha(request, fornecedor_id=None):
    
    if fornecedor_id:
         try:
             fornecedor = Fornecedores.objects.get(id=fornecedor_id)
         except Fornecedores.DoesNotExist:
             messages.error(request, "Fornecedor não encontrado.")
             return redirect('fornecedores')
    else:
         fornecedor = None
    
    #salvar
    if request.method == 'POST':
        if fornecedor:
            fornecedor_form = FornecedoresForm(request.POST, instance=fornecedor)
            novo_fornecedor = False
        else:
            fornecedor_form = FornecedoresForm(request.POST)
            novo_fornecedor = True

        #Conferir campos obrigatórios
        fields = [
            ('cnpj', "O CNPJ é obrigatório!"),
            ('razao_social', "A razão social é obrigatória!"),
            ('porte', "O porte é obrigatório!"),
            ('tipo_direito', "O tipo de direito é obrigatório!"),
            ('end_uf', 'A UF é obrigatória!')
        ]
        for field_name, error_message in fields:
            valor = request.POST.get(field_name)
            if valor == "Não informado" or valor == None:
                messages.error(request, error_message)
                if fornecedor:
                    return JsonResponse({
                    'redirect_url': reverse('fornecedor_ficha', args=[fornecedor.id]),
                    })
                else:
                    return JsonResponse({
                    'redirect_url': reverse('fornecedor_novo'),
                    'data': request.POST,
                    })

        #Verificar se houve alteração no formulário
        if not fornecedor_form.has_changed():
            messages.error(request, "Dados não foram salvos. Não houve mudanças.")
            if fornecedor:
                return JsonResponse({
                    'redirect_url': reverse('fornecedor_ficha', args=[fornecedor.id]),
                })
            else:
                #return redirect('fornecedor_novo')
                return JsonResponse({
                    'redirect_url': reverse('fornecedor_novo'),
                })

        if fornecedor_form.is_valid():
            #Verificar se já existe o fornecedor na base
            cnpj = fornecedor_form.cleaned_data.get('cnpj')
            cnpj_existente = Fornecedores.objects.filter(cnpj=cnpj)
            
            #Se estivermos atualizando um fornecedor existente, excluímos esse fornecedor da verificação
            if fornecedor:
                cnpj_existente = cnpj_existente.exclude(id=fornecedor.id)

            if cnpj_existente.exists():
                messages.error(request, "Já existe um fornecedor com esse CNPJ. Não foi possível salvar.")
                if fornecedor:
                    return JsonResponse({
                        'redirect_url': reverse('fornecedor_ficha', args=[fornecedor.id]),
                    })
                else:
                    return JsonResponse({
                        'redirect_url': reverse('fornecedor_novo'),
                    })

            #Salvar o produto
            fornecedor = fornecedor_form.save(commit=False)
            fornecedor.save(current_user=request.user.usuario_relacionado)
            if novo_fornecedor:
                messages.success(request, "Novo fornecedor registrado com sucesso!")
            else:
                messages.success(request, "Dados atualizados com sucesso!")
            
            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Fornecedores_Fornecedores",
                model='Fornecedores',
                model_id={fornecedor.id},
                item_id=0,
                item_descricao="Salvar edição de fornecedor.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o fornecedor {fornecedor.nome_fantasia} ({fornecedor.cnpj}) em {current_date_str}."
            )
            log_entry.save()
            
            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'redirect_url': reverse('fornecedor_ficha', args=[fornecedor.id]),
                })
   
        else:
            messages.error(request, "Formulário inválido")
            print("Erro formulário fornecedor")
            print(fornecedor_form.errors)

    #novo
    form = FornecedoresForm(instance=fornecedor)
    lista_naturezajuridica = CNPJ_NATUREZA_JURIDICA.objects.values_list('codigo', flat=True).distinct().order_by('codigo')
    lista_cnae = CNPJ_CNAE.objects.values_list('subclasse_codigo', flat=True).distinct().order_by('subclasse_codigo')
    lista_ufs = UF_Municipio.objects.values_list('uf_sigla', flat=True).distinct().order_by('uf_sigla')
    uf = form.instance.end_uf
    lista_municipios = UF_Municipio.objects.filter(uf_sigla=uf).values_list('municipio', flat=True).order_by('municipio')
    
    return render(request, 'fornecedores/fornecedor_ficha.html', {
        'fornecedor': fornecedor,
        'form': form,
        'lista_cnpj_hierarquia': CNPJ_HIERARQUIA,
        'lista_cnpj_porte': CNPJ_PORTE,
        'lista_tipo_direito': TIPO_DIREITO,
        'lista_ufs': lista_ufs,
        'lista_naturezajuridica': lista_naturezajuridica,
        'lista_cnae': lista_cnae,
        'lista_municipios': lista_municipios,
    })

def fornecedor_ficha_filtrar_dados(request):
    tipo = request.GET.get('tipo')
    
    if tipo == 'natjuridica':
        codigo = request.GET.get('codigo')
        nat_juridica = CNPJ_NATUREZA_JURIDICA.objects.get(codigo=codigo)
        return JsonResponse({'valor': nat_juridica.natureza_juridica})

    elif tipo == 'ativ_principal':
        codigo = request.GET.get('codigo')
        atividade_principal = CNPJ_CNAE.objects.get(subclasse_codigo=codigo)
        return JsonResponse({'valor': atividade_principal.subclasse_descricao})

    elif tipo == 'end_uf':
        uf = request.GET.get('uf')
        municipios = UF_Municipio.objects.filter(uf_sigla=uf).values('id', 'municipio')
        return JsonResponse({'municipios': list(municipios)})

    else:
        return JsonResponse({'error': 'Tipo não reconhecido'}, status=400)

def fornecedor_delete(request, fornecedor_id=None):
    try:
        fornecedor = Fornecedores.objects.get(id=fornecedor_id)
        fornecedor.soft_delete(request.user.usuario_relacionado)
        messages.error(request, "Fornecedor deletado com sucesso.")

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Fornecedores_Fornecedores",
            model='Fornecedores',
            model_id={fornecedor.id},
            item_id=0,
            item_descricao="Deleção de fornecedor.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o fornecedor {fornecedor.nome_fantasia} ({fornecedor.cnpj}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({"message": "Fornecedor deletado com sucesso!"})
    except Fornecedores.DoesNotExist:
        messages.error(request, "Fornecedor não encontrado.")    
    return redirect('fornecedores')

def fornecedores_filtro(request):
    hierarquia = request.GET.get('hierarquia', None)
    cnpj_porte = request.GET.get('cnpj_porte', None)
    tipo_direito = request.GET.get('tipo_direito', None)
    uf_fornecedor = request.GET.get('uf_fornecedor', None)
    fornecedor = request.GET.get('fornecedor', None)

    filters = {}
    filters['del_status'] = False
    if hierarquia:
        filters['hierarquia'] = hierarquia
    if cnpj_porte:
        filters['porte'] = cnpj_porte
    if tipo_direito:
        filters['tipo_direito'] = tipo_direito
    if uf_fornecedor:
        filters['end_uf'] = uf_fornecedor
    if fornecedor:
        filters['nome_fantasia__icontains'] = fornecedor
    
    tab_fornecedores = Fornecedores.objects.filter(**filters).order_by('nome_fantasia')
    total_fornecedores = tab_fornecedores.count()
    
    page = int(request.GET.get('page', 1))
    paginator = Paginator(tab_fornecedores, 100)  # Mostra 100 fornecedores por página
    try:
        fornecedores_paginados = paginator.page(page)
    except EmptyPage:
        fornecedores_paginados = paginator.page(paginator.num_pages)

    data = list(fornecedores_paginados.object_list.values())
    
    return JsonResponse({
        'data': data,
        'total_fornecedores': total_fornecedores,
        'has_next': fornecedores_paginados.has_next(),
        'has_previous': fornecedores_paginados.has_previous(),
        'current_page': page
    })

def fornecedores_exportar(request):
    print("Exportar Fornecedores")
    
    if request.method == 'POST':
        data = json.loads(request.body)
        hierarquia = data.get('hierarquia')
        cnpj_porte = data.get('cnpj_porte')
        tipo_direito = data.get('tipo_direito')
        uf_fornecedor = data.get('uf_fornecedor')
        fornecedor = data.get('fornecedor_nome')
    
        print('Hierarquia: ', hierarquia)
        print(cnpj_porte)
        print(tipo_direito)
        print(uf_fornecedor)
        print(fornecedor)

        filters = {}
        filters['del_status'] = False
        if hierarquia:
            filters['hierarquia'] = hierarquia
        if cnpj_porte:
            filters['porte'] = cnpj_porte
        if tipo_direito:
            filters['tipo_direito'] = tipo_direito
        if uf_fornecedor:
            filters['end_uf'] = uf_fornecedor
        if fornecedor:
            filters['nome_fantasia__icontains'] = fornecedor
        
        fornecedores = Fornecedores.objects.filter(**filters)
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "fornecedores"

        headers = [
        'ID', 'Usuário Registro', 'Usuário Atualização', 'Data de Registro', 'Data da Última Atualização',
        'N Edições', 'CNPJ', 'Razão Social', 'Nome Fantasia', 'Hierarquia', 'Porte', 'Tipo de Direito',
        'Data de Abertura', 'Código NatJud', 'Natureza Jurídica', 'Código AtivPri', 'Atividade Principal',
        'CEP', 'UF', 'Município', 'Logradouro', 'Número', 'Bairro', 'Observações Gerais', 'Data Exportação'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 25

        # Adicionar dados da tabela
        for row_num, fornecedor in enumerate(fornecedores, 2):
            ws.cell(row=row_num, column=1, value=fornecedor.id)
            ws.cell(row=row_num, column=2, value=str(fornecedor.usuario_registro.primeiro_ultimo_nome()))
            ws.cell(row=row_num, column=3, value=str(fornecedor.usuario_atualizacao.primeiro_ultimo_nome()))
            registro_data = fornecedor.registro_data.replace(tzinfo=None)
            ult_atual_data = fornecedor.ult_atual_data.replace(tzinfo=None)
            ws.cell(row=row_num, column=4, value=registro_data)
            ws.cell(row=row_num, column=5, value=ult_atual_data)
            ws.cell(row=row_num, column=6, value=fornecedor.log_n_edicoes)
            ws.cell(row=row_num, column=7, value=fornecedor.cnpj)
            ws.cell(row=row_num, column=8, value=fornecedor.razao_social)
            ws.cell(row=row_num, column=9, value=fornecedor.nome_fantasia)
            ws.cell(row=row_num, column=10, value=fornecedor.hierarquia)
            ws.cell(row=row_num, column=11, value=fornecedor.porte)
            ws.cell(row=row_num, column=12, value=fornecedor.tipo_direito)
            ws.cell(row=row_num, column=13, value=fornecedor.data_abertura)
            ws.cell(row=row_num, column=14, value=fornecedor.natjuridica_codigo)
            ws.cell(row=row_num, column=15, value=fornecedor.natjuridica_descricao)
            ws.cell(row=row_num, column=16, value=fornecedor.ativ_principal_cod)
            ws.cell(row=row_num, column=17, value=fornecedor.ativ_principal_descricao)
            ws.cell(row=row_num, column=18, value=fornecedor.end_cep)
            ws.cell(row=row_num, column=19, value=fornecedor.end_uf)
            ws.cell(row=row_num, column=20, value=fornecedor.end_municipio)
            ws.cell(row=row_num, column=21, value=fornecedor.end_logradouro)
            ws.cell(row=row_num, column=22, value=fornecedor.end_numero)
            ws.cell(row=row_num, column=23, value=fornecedor.end_bairro)
            ws.cell(row=row_num, column=24, value=fornecedor.observacoes_gerais)
            ws.cell(row=row_num, column=25, value=current_date_str)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Reposition to the start of the stream

        # Registrar a ação no CustomLog
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Fornecedores",
            model='Fornecedores',
            model_id=0,   
            item_id=0,
            item_descricao="Exportação da lista de fornecedores",
            acao="Exportação",
            observacoes=f"Usuário {request.user.username} exportou lista de fornecedores em {current_date_str}."
        )
        log_entry.save()

        # Configurar a resposta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportar_fornecedores.xlsx"'
        response.write(output.getvalue())
        return response
    
def fornecedores_buscar(request):
    fornecedores = Fornecedores.objects.filter(del_status=False).order_by('nome_fantasia')
    fornecedores_list = list(fornecedores.values('id', 'cnpj', 'nome_fantasia'))
    return JsonResponse({'fornecedores_list': fornecedores_list})



#FORNECEDORES FAQs
def fornecedores_faq(request):
    tab_fornecedores_faq = Fornecedores_Faq.objects.filter(del_status=False).order_by('topico')
    total_fornecedores_faq = tab_fornecedores_faq.count()
    conteudo = {
         'tab_fornecedores_faq': tab_fornecedores_faq,
         'total_fornecedores_faq': total_fornecedores_faq,
         'lista_topico': FAQ_FORNECEDOR_TOPICO,
    }
    return render(request, 'fornecedores/fornecedores_faq.html', conteudo)

def fornecedor_faq_ficha(request, faq_id=None):
    if faq_id:
         try:
             faq = Fornecedores_Faq.objects.get(id=faq_id)
         except Fornecedores_Faq.DoesNotExist:
             messages.error(request, "FAQ não encontrado.")
             return redirect('fornecedores_faq')
    else:
         faq = None
    
    #salvar
    if request.method == 'POST':
        
        if faq:
            faq_form = FornecedoresFaqForm(request.POST, instance=faq)
            novo_faq = False
        else:
            faq_form = FornecedoresFaqForm(request.POST)
            novo_faq = True

        #Conferir campos obrigatórios
        fields = [
            ('topico', "O Tópico/Assunto é obrigatório!"),
            ('contexto', "O Contexto é obrigatório!"),
            ('resposta', "A Resposta é obrigatória!"),
        ]
        for field_name, error_message in fields:
            valor = request.POST.get(field_name)
            print('valor= ', valor)
            if valor == "Não informado" or valor == None or valor == '':
                messages.error(request, error_message)
                if faq:
                    return JsonResponse({
                        'redirect_url': reverse('fornecedor_faq_ficha', args=[faq.id]),
                    })
                else:
                    return JsonResponse({
                        'redirect_url': reverse('fornecedor_faq_novo'),
                        'data': request.POST,
                    })

        #Verificar se houve alteração no formulário
        if not faq_form.has_changed():
            messages.error(request, "Dados não foram salvos. Não houve mudanças.")
            if faq:
                return JsonResponse({
                    'redirect_url': reverse('fornecedor_faq_ficha', args=[faq.id]),
                })
            else:
                #return redirect('fornecedor_novo')
                return JsonResponse({
                    'redirect_url': reverse('fornecedor_faq_novo'),
                })

        if faq_form.is_valid():
            #Salvar o produto
            faq = faq_form.save(commit=False)
            faq.save(current_user=request.user.usuario_relacionado)
            if novo_faq:
                messages.success(request, "Novo FAQ registrado com sucesso!")
            else:
                messages.success(request, "Dados atualizados com sucesso!")
            
            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Fornecedores_FAQs",
                model='Fornecedores_Faq',
                model_id={faq.id},
                item_id=0,
                item_descricao="Salvar edição de FAQ.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o FAQ (ID {faq.id}, Tópico: {faq.topico}, Contexto: {faq.contexto}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'redirect_url': reverse('fornecedor_faq_ficha', args=[faq.id]),
                })
        else:
            messages.error(request, "Formulário inválido")
            print("Erro formulário FAQ")
            print(faq_form.errors)

    #novo
    form = FornecedoresFaqForm(instance=faq)
    
    return render(request, 'fornecedores/fornecedor_faq_ficha.html', {
        'lista_topico': FAQ_FORNECEDOR_TOPICO,
        'fornecedor_faq': faq,
        'form': form,
    })

def fornecedor_faq_filtrar_dados(request):
    topico = request.GET.get('topico', None)
    contexto = request.GET.get('contexto', None)
    resposta = request.GET.get('resposta', None)
    print('Topico: ', topico)
    filters = {}
    filters['del_status'] = False
    if topico:
        filters['topico'] = topico
    if contexto:
        filters['contexto__icontains'] = contexto
    if resposta:
        filters['resposta__icontains'] = resposta
    
    tab_fornecedores_faq = Fornecedores_Faq.objects.filter(**filters).order_by('topico')
    total_fornecedores_faq = tab_fornecedores_faq.count()
    
    page = int(request.GET.get('page', 1))
    paginator = Paginator(tab_fornecedores_faq, 100)  # Mostra 100 faqs por página
    try:
        faq_paginados = paginator.page(page)
    except EmptyPage:
        faq_paginados = paginator.page(paginator.num_pages)

    data = list(faq_paginados.object_list.values())
    
    return JsonResponse({
        'data': data,
        'total_fornecedores_faq': total_fornecedores_faq,
        'has_next': faq_paginados.has_next(),
        'has_previous': faq_paginados.has_previous(),
        'current_page': page
    })

def fornecedor_faq_delete(request, faq_id=None):
    try:
        faq = Fornecedores_Faq.objects.get(id=faq_id)
        faq.soft_delete(request.user.usuario_relacionado)
        messages.error(request, "FAQ deletado com sucesso.")

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Fornecedores_FAQs",
            model='Fornecedores_Faq',
            model_id={faq.id},
            item_id=0,
            item_descricao="Deleção de FAQ.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o FAQ (ID {faq.id}, Tópico: {faq.topico}, Contexto: {faq.contexto}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({"message": "Fornecedor deletado com sucesso!"})
    except Fornecedores_Faq.DoesNotExist:
        messages.error(request, "FAQ não encontrado.")    
    return redirect('fornecedores_faq')

def fornecedores_faq_exportar(request):
    print("Exportar FAQs Fornecedores")
    
    if request.method == 'POST':
        data = json.loads(request.body)
        topico = data.get('topico')
        contexto = data.get('contexto')
        resposta = data.get('resposta')

        filters = {}
        filters['del_status'] = False
        if topico:
            filters['topico'] = topico
        if contexto:
            filters['contexto__icontains'] = contexto
        if resposta:
            filters['resposta__icontains'] = resposta
        
        faqs = Fornecedores_Faq.objects.filter(**filters)
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "fornecedores_faqs"

        headers = [
        'ID', 'Usuário Registro', 'Usuário Atualização', 'Data de Registro', 'Data da Última Atualização',
        'N Edições', 'Tópico', 'Outro Tópico', 'Contexto', 'Resposta', 'Observações Gerais', 'Data Exportação'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 25

        # Adicionar dados da tabela
        for row_num, faq in enumerate(faqs, 2):
            ws.cell(row=row_num, column=1, value=faq.id)
            ws.cell(row=row_num, column=2, value=str(faq.usuario_registro.primeiro_ultimo_nome()))
            ws.cell(row=row_num, column=3, value=str(faq.usuario_atualizacao.primeiro_ultimo_nome()))
            registro_data = faq.registro_data.replace(tzinfo=None)
            ult_atual_data = faq.ult_atual_data.replace(tzinfo=None)
            ws.cell(row=row_num, column=4, value=registro_data)
            ws.cell(row=row_num, column=5, value=ult_atual_data)
            ws.cell(row=row_num, column=6, value=faq.log_n_edicoes)
            ws.cell(row=row_num, column=7, value=faq.topico)
            ws.cell(row=row_num, column=8, value=faq.topico_outro)
            ws.cell(row=row_num, column=9, value=faq.contexto)
            ws.cell(row=row_num, column=10, value=faq.resposta)
            ws.cell(row=row_num, column=11, value=faq.observacoes_gerais)
            ws.cell(row=row_num, column=12, value=current_date_str)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Reposition to the start of the stream

        # Registrar a ação no CustomLog
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Fornecedores_FAQs",
            model='Fornecedores_Faq',
            model_id=0,
            item_id=0,
            item_descricao="Exportação da lista de faqs dos fornecedores",
            acao="Exportação",
            observacoes=f"Usuário {request.user.username} exportou lista de faqs dos fornecedores em {current_date_str}."
        )
        log_entry.save()

        # Configurar a resposta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportar_fornecedores.xlsx"'
        response.write(output.getvalue())
        return response





#FORNECEDORES REPRESENTANTES
def fornecedores_representantes(request, id_fornecedor=None):

    if request.method == 'POST':
        
        id_representante = request.POST.get('id_representante')

        if id_representante:
            try:
                representante = Fornecedores_Representantes.objects.get(id=id_representante)
            except Fornecedores_Representantes.DoesNotExist:
                messages.error(request, "Representante não encontrado.")
                return redirect('fornecedores')
        else:
            representante = None

        if representante:
            representante_form = FornecedoresRepresentantesForm(request.POST, instance=representante)
            novo_representante = False
        else:
            representante_form = FornecedoresRepresentantesForm(request.POST)
            novo_representante = True
        
        #Verificar se houve alteração no formulário
        if not representante_form.has_changed():
            messages.error(request, "Dados não foram salvos. Não houve mudanças.")
            return redirect('fornecedores_representantes', id_fornecedor=id_fornecedor)
        
        fornecedor_instance = Fornecedores.objects.get(id=id_fornecedor)
        representante_form.instance.fornecedor = fornecedor_instance
        
        if representante_form.is_valid():
            #Verificar se já existe registro dessa tramitacao
            cpf = representante_form.cleaned_data.get('cpf')
            cpf_existente = Fornecedores_Representantes.objects.filter(cpf=cpf, fornecedor=fornecedor_instance)

            #Se estivermos atualizando um processo existente, excluímos esse processo da verificação
            # if cpf:
            #     if representante:
            #         cpf_existente = cpf_existente.exclude(id=representante.id)
                
            #     if cpf_existente.exists():
            #         messages.error(request, "Já existe um representante com esse CPF. Não foi possível salvar.")
            #         return redirect('fornecedores_representantes', id_fornecedor=id_fornecedor)
            
            #Salvar a tramitação
            representante = representante_form.save(commit=False)
            representante.save(current_user=request.user.usuario_relacionado)
            representante_id = representante.id

            if novo_representante:
                messages.success(request, "Novo representante registrado com sucesso!")
            else:
                messages.success(request, "Dados atualizados com sucesso!")

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Fornecedores_Fornecedor_Representantes",
                model='Fornecedores_Representantes',
                model_id={representante.id},
                item_id=0,
                item_descricao="Salvar edição de Representante do Fornecedor.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o Representante (ID {representante.id}, Nome: {representante.nome_completo}, Fornecedor: {representante.fornecedor.cnpj}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar log
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'redirect_url': reverse('fornecedores_representantes', args=[id_fornecedor]),
                    'representante_id': representante_id,
                })
        else:
            messages.error(request, "Preencha os campos obrigatórios.")
            print("Erro formulário representante do fornecedor.")
            print(representante_form.errors)
    
    if id_fornecedor:
        fornecedor_instance = Fornecedores.objects.get(id=id_fornecedor)
        tab_fornecedores_representantes = Fornecedores_Representantes.objects.filter(del_status=False, fornecedor=fornecedor_instance).order_by('nome_completo')
    else:
        tab_fornecedores_representantes = None

    conteudo = {
        'tab_fornecedores_representantes': tab_fornecedores_representantes,
        'lista_cargos': CARGOS_FUNCOES,
        'lista_genero_sexual': GENERO_SEXUAL,
        'id_fornecedor': id_fornecedor,
    }
    return render(request, 'fornecedores/fornecedor_representantes.html', conteudo)

def fornecedor_representante_delete(request, representante_id=None):
    try:
        representante = Fornecedores_Representantes.objects.get(id=representante_id)
        representante.soft_delete(request.user.usuario_relacionado)
        messages.error(request, "Representante deletado com sucesso.")

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Fornecedores_Fornecedor_Representantes",
            model='Fornecedores_Representantes',
            model_id={representante.id},
            item_id=0,
            item_descricao="Deleção de Representante do Fornecedor.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o Representante (ID {representante.id}, Nome: {representante.nome_completo}, Fornecedor: {representante.fornecedor.cnpj}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({"message": "Representante deletado com sucesso!"})
    except Fornecedores_Representantes.DoesNotExist:
        messages.error(request, "Representante não encontrado.")
        return JsonResponse({"message": "Representante não encontrado."})   

def representante_dados(request, representante_id):
    try:
        representante = Fornecedores_Representantes.objects.get(id=representante_id)
        print('Data do Registro ', representante.registro_data)
        data = {
            'id_representante': representante.id,
            'log_data_registro': representante.registro_data.strftime('%d/%m/%Y %H:%M:%S') if representante.registro_data else '',
            'log_responsavel_registro': str(representante.usuario_atualizacao.dp_nome_completo),
            'lot_ult_atualizacao': representante.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if representante.ult_atual_data else '',
            'log_responsavel_atualizacao': str(representante.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': representante.log_n_edicoes,
            'cpf': representante.cpf if representante.cpf else '',
            'nome_completo': representante.nome_completo,
            'data_nascimento': representante.data_nascimento.strftime('%d/%m/%Y') if representante.data_nascimento else '',
            'genero_sexual': representante.genero_sexual,
            'cargo': representante.cargo,
            'cargo_outro': representante.cargo_outro,
            'telefone': representante.telefone,
            'celular': representante.celular,
            'email': representante.email,
            'linkedin': representante.linkedin,
            'observacoes': representante.observacoes_gerais if representante.observacoes_gerais else '',
        }
        return JsonResponse(data)
    except Fornecedores_Representantes.DoesNotExist:
        return JsonResponse({'error': 'Representante não encontrado'}, status=404)

def fornecedor_representantes_exportar(request, id_fornecedor):
    data = json.loads(request.body)

    #fornecedor = Fornecedores.objects.get(id=fornecedor_id)
    print('Fornecedor ', id_fornecedor)
    filters = {}
    filters['del_status'] = False
    filters['fornecedor_id'] = id_fornecedor
    
    representantes = Fornecedores_Representantes.objects.filter(**filters)
    current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    # Criar um workbook e adicionar uma planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "fornecedores_comunicacoes"

    headers = [
    'ID', 'Usuário Registro', 'Usuário Atualização', 'Data de Registro', 'Data da Última Atualização', 'N Edições',
    'CNPJ', 'Fornecedor',
    'CPF', 'Nome Completo', 'Data de Nascimento', 'Gênero Sexual',
    'Cargo/Função', 'Outro Cargo/Função', 'Telefone', 'Celular', 'Email', 'LinkedIn',
    'Observações Gerais', 'Data Exportação'
    ]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = header
        ws.column_dimensions[col_letter].width = 20

    # Adicionar dados da tabela
    for row_num, item in enumerate(representantes, 2):
        registro_data = item.registro_data.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M:%S')
        ult_atual_data = item.ult_atual_data.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M:%S')
        data_nascimento = item.data_nascimento
        if data_nascimento:
            data_nascimento.strftime('%d/%m/%Y')
        fornecedor = item.fornecedor.nome_fantasia
        cnpj = item.fornecedor.cnpj

        ws.cell(row=row_num, column=1, value=item.id)
        ws.cell(row=row_num, column=2, value=str(item.usuario_registro.primeiro_ultimo_nome()))
        ws.cell(row=row_num, column=3, value=str(item.usuario_atualizacao.primeiro_ultimo_nome()))
        ws.cell(row=row_num, column=4, value=registro_data)
        ws.cell(row=row_num, column=5, value=ult_atual_data)
        ws.cell(row=row_num, column=6, value=item.log_n_edicoes)

        ws.cell(row=row_num, column=7, value=cnpj)
        ws.cell(row=row_num, column=8, value=fornecedor)
        ws.cell(row=row_num, column=9, value=item.cpf)
        ws.cell(row=row_num, column=10, value=item.nome_completo)
        ws.cell(row=row_num, column=11, value=data_nascimento)
        ws.cell(row=row_num, column=12, value=item.genero_sexual)

        ws.cell(row=row_num, column=13, value=item.cargo)
        ws.cell(row=row_num, column=14, value=item.cargo_outro)
        ws.cell(row=row_num, column=15, value=item.telefone)
        ws.cell(row=row_num, column=16, value=item.celular)
        ws.cell(row=row_num, column=17, value=item.email)
        ws.cell(row=row_num, column=18, value=item.linkedin)

        ws.cell(row=row_num, column=19, value=item.observacoes_gerais)
        ws.cell(row=row_num, column=20, value=current_date_str)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Reposition to the start of the stream

    # Registrar a ação no CustomLog
    log_entry = CustomLog(
        usuario=request.user.usuario_relacionado,
        modulo="Fornecedores_Representantes",
        model='Fornecedores_Representantes',
        model_id=0,
        item_id=0,
        item_descricao="Exportação dos dados dos representantes do fornecedor.",
        acao="Exportação",
        observacoes=f"Usuário {request.user.username} exportou dados dos representantes do fornecedor {fornecedor} ({cnpj}) em {current_date_str}."
    )
    log_entry.save()

    # Configurar a resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="exportar_fornecedores.xlsx"'
    response.write(output.getvalue())
    return response





#FORNECEDORES COMUNICACOES
def fornecedores_comunicacoes(request, id_fornecedor=None):

    if request.method == 'POST':
        
        id_comunicacao = request.POST.get('id_comunicacao')

        if id_comunicacao:
            try:
                comunicacao = Fornecedores_Comunicacoes.objects.get(id=id_comunicacao)
            except Fornecedores_Comunicacoes.DoesNotExist:
                messages.error(request, "Comunicação não encontrada.")
                return redirect('fornecedores')
        else:
            comunicacao = None

        if comunicacao:
            comunicacao_form = FornecedoresComunicacoesForm(request.POST, instance=comunicacao)
            nova_comunicacao = False
        else:
            comunicacao_form = FornecedoresComunicacoesForm(request.POST)
            nova_comunicacao = True
        
        #Verificar se houve alteração no formulário
        if not comunicacao_form.has_changed():
            messages.error(request, "Dados não foram salvos. Não houve mudanças.")
            return redirect('fornecedores_comunicacoes', id_fornecedor=id_fornecedor)
        
        fornecedor_instance = Fornecedores.objects.get(id=id_fornecedor)
        comunicacao_form.instance.fornecedor = fornecedor_instance
        
        # #Responsavel
        responsavel = request.POST.get('responsavel_resposta')
        print(responsavel)
        if responsavel == 'outro':
            #comunicacao_form.instance.responsavel_resposta = responsavel
            comunicacao_form.instance.responsavel_resposta = None

        if comunicacao_form.is_valid():

            #Salvar a comunicacao
            comunicacao = comunicacao_form.save(commit=False)
            comunicacao.save(current_user=request.user.usuario_relacionado)
            comunicacao_id = comunicacao.id

            if nova_comunicacao:
                messages.success(request, "Nova comunicação registrada com sucesso!")
            else:
                messages.success(request, "Dados atualizados com sucesso!")

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Fornecedores_Comunicação",
                model='Fornecedores_Comunicacoes',
                model_id={comunicacao.id},
                item_id=0,
                item_descricao="Salvar edição de comunicação com o fornecedor.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou a comunicacao (ID: {comunicacao.id}) com o fornecedor {comunicacao.fornecedor.nome_fantasia} (CNPJ: {comunicacao.fornecedor.cnpj}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar log
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'redirect_url': reverse('fornecedores_comunicacoes', args=[id_fornecedor]),
                    'comunicacao_id': comunicacao_id,
                })
        else:
            messages.error(request, "Preencha os campos obrigatórios.")
            print("Erro formulário representante do fornecedor.")
            print(comunicacao_form.errors)
    
    if id_fornecedor:
        fornecedor_instance = Fornecedores.objects.get(id=id_fornecedor)
        tab_fornecedores_comunicacoes = Fornecedores_Comunicacoes.objects.filter(del_status=False, fornecedor=fornecedor_instance).order_by('-data_envio')
    else:
        tab_fornecedores_comunicacoes = None

    conteudo = {
        'tab_fornecedores_comunicacoes': tab_fornecedores_comunicacoes,
        'lista_topicos': FAQ_FORNECEDOR_TOPICO,
        'lista_tipo_comunicacao': TIPO_COMUNICACAO,
        'lista_status_envio': STATUS_ENVIO_COMUNICACAO,
        'lista_unidade_daf': UNIDADE_DAF2,
        'id_fornecedor': id_fornecedor,
    }
    return render(request, 'fornecedores/fornecedor_comunicacoes.html', conteudo)

def fornecedor_comunicacao_delete(request, comunicacao_id=None):
    try:
        comunicacao = Fornecedores_Comunicacoes.objects.get(id=comunicacao_id)
        comunicacao.soft_delete(request.user.usuario_relacionado)
        messages.error(request, "Comunicação deletada com sucesso.")

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Fornecedores_Comunicação",
            model='Fornecedores_Comunicacoes',
            model_id={comunicacao.id},
            item_id=0,
            item_descricao="Deleção de comunicação com o fornecedor.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou a comunicacao (ID: {comunicacao.id}) com o fornecedor {comunicacao.fornecedor.nome_fantasia} (CNPJ: {comunicacao.fornecedor.cnpj}) em {current_date_str}."
        )
        log_entry.save()
        
        return JsonResponse({"message": "Comunicação deletada com sucesso!"})
    except Fornecedores_Comunicacoes.DoesNotExist:
        messages.error(request, "Comunicação não encontrada.")
        return JsonResponse({"message": "Comunicação não encontrada."})   

def comunicacao_dados(request, comunicacao_id):
    try:
        item = Fornecedores_Comunicacoes.objects.get(id=comunicacao_id)
        if item.responsavel_resposta is None:
            if not item.outro_responsavel:
                responsavel_resposta = "Não Informado"
            else:
                responsavel_resposta = "Outro"
        else:
            responsavel_resposta = str(item.responsavel_resposta.dp_nome_completo)
        data = {
            'id_comunicacao': item.id,
            'log_data_registro': item.registro_data.strftime('%d/%m/%Y %H:%M:%S') if item.registro_data else '',
            'log_responsavel_registro': str(item.usuario_atualizacao.dp_nome_completo),
            'lot_ult_atualizacao': item.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if item.ult_atual_data else '',
            'log_responsavel_atualizacao': str(item.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': item.log_n_edicoes,
            'unidade_daf': item.unidade_daf,
            'tipo_comunicacao': item.tipo_comunicacao,
            'topico_comunicacao': item.topico_comunicacao,
            'assunto': item.assunto,
            'demanda_original': item.demanda_original,
            'destinatario': item.destinatario,
            'mensagem_encaminhada': item.mensagem_encaminhada,
            'status_envio': item.status_envio,
            'data_envio': item.data_envio.strftime('%Y-%m-%d') if item.data_envio else '',
            'responsavel_resposta': responsavel_resposta,
            'outro_responsavel': item.outro_responsavel,
            'observacoes': item.observacoes_gerais if item.observacoes_gerais else '',
        }
        return JsonResponse(data)
    except Fornecedores_Comunicacoes.DoesNotExist:
        return JsonResponse({'error': 'Comunicação não encontrada'}, status=404)

def fornecedor_usuarios_por_unidade(request, unidade):
    usuarios = Usuarios.usuarios_por_unidade(unidade)
    return JsonResponse(usuarios, safe=False)

def fornecedor_comunicacao_exportar(request, id_fornecedor):
    data = json.loads(request.body)

    #fornecedor = Fornecedores.objects.get(id=fornecedor_id)
    print('Fornecedor ', id_fornecedor)
    filters = {}
    filters['del_status'] = False
    filters['fornecedor_id'] = id_fornecedor
    
    comunicacoes = Fornecedores_Comunicacoes.objects.filter(**filters)
    current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    # Criar um workbook e adicionar uma planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "fornecedores_comunicacoes"

    headers = [
    'ID', 'Usuário Registro', 'Usuário Atualização', 'Data de Registro', 'Data da Última Atualização', 'N Edições',
    'CNPJ', 'Fornecedor',
    'Unidade DAF', 'Tipo de Comunicação', 'Tópico da Comunicação', 'Assunto', 
    'Demanda Original', 'Destinatário', 'Mensagem Encaminhada',
    'Status do Envio', 'Data do Envio', 'Responsável pela Resposta', 'Outro Responsável',
    'Observações Gerais', 'Data Exportação'
    ]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = header
        ws.column_dimensions[col_letter].width = 20

    # Adicionar dados da tabela
    for row_num, item in enumerate(comunicacoes, 2):
        registro_data = item.registro_data.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M:%S')
        ult_atual_data = item.ult_atual_data.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M:%S')
        data_envio = item.data_envio
        if data_envio:
            data_envio.strftime('%d/%m/%Y')
        responsavel_resposta = item.responsavel_resposta
        outro_responsavel = item.outro_responsavel
        if item.responsavel_resposta:
            responsavel_resposta = item.responsavel_resposta.dp_nome_completo
        elif item.responsavel_resposta == None and outro_responsavel != '':
            responsavel_resposta == 'Outro'
        else:
            responsavel_resposta == 'Não Informado'
        fornecedor = item.fornecedor.nome_fantasia
        cnpj = item.fornecedor.cnpj

        ws.cell(row=row_num, column=1, value=item.id)
        ws.cell(row=row_num, column=2, value=str(item.usuario_registro.primeiro_ultimo_nome()))
        ws.cell(row=row_num, column=3, value=str(item.usuario_atualizacao.primeiro_ultimo_nome()))
        ws.cell(row=row_num, column=4, value=registro_data)
        ws.cell(row=row_num, column=5, value=ult_atual_data)
        ws.cell(row=row_num, column=6, value=item.log_n_edicoes)

        ws.cell(row=row_num, column=7, value=cnpj)
        ws.cell(row=row_num, column=8, value=fornecedor)
        ws.cell(row=row_num, column=9, value=item.unidade_daf)
        ws.cell(row=row_num, column=10, value=item.tipo_comunicacao)
        ws.cell(row=row_num, column=11, value=item.topico_comunicacao)
        ws.cell(row=row_num, column=12, value=item.assunto)

        ws.cell(row=row_num, column=13, value=item.demanda_original)
        ws.cell(row=row_num, column=14, value=item.destinatario)
        ws.cell(row=row_num, column=15, value=item.mensagem_encaminhada)

        ws.cell(row=row_num, column=16, value=item.status_envio)
        ws.cell(row=row_num, column=17, value=data_envio)
        ws.cell(row=row_num, column=18, value=responsavel_resposta)
        ws.cell(row=row_num, column=19, value=outro_responsavel)

        ws.cell(row=row_num, column=20, value=item.observacoes_gerais)
        ws.cell(row=row_num, column=21, value=current_date_str)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Reposition to the start of the stream

    # Registrar a ação no CustomLog
    log_entry = CustomLog(
        usuario=request.user.usuario_relacionado,
        modulo="Fornecedores_Comunicação",
        model='Fornecedores_Comunicacoes',
        model_id=0,
        item_id=0,
        item_descricao="Exportação dos registros de comunicação do fornecedor.",
        acao="Exportação",
        observacoes=f"Usuário {request.user.username} exportou registros de comunicações com o fornecedor {fornecedor} ({cnpj}) em {current_date_str}."
    )
    log_entry.save()

    # Configurar a resposta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="exportar_fornecedores.xlsx"'
    response.write(output.getvalue())
    return response