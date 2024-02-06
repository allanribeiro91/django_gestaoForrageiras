from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator, EmptyPage
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import auth, messages
from django.contrib.auth.models import User
from apps.usuarios.models import Usuarios
from apps.produtos.models import DenominacoesGenericas, ProdutosFarmaceuticos, Tags, ListaATC, ProdutosTags, ProdutoConsumoMedio
from apps.produtos.forms import DenominacoesGenericasForm, ProdutosFarmaceuticosForm
from apps.main.models import CustomLog
from setup.choices import TIPO_PRODUTO, FORMA_FARMACEUTICA, STATUS_INCORPORACAO, CONCENTRACAO_TIPO, YES_NO, CLASSIFICACAO_AWARE
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import json

@login_required
def produtos(request):
    produtos = ProdutosFarmaceuticos.objects.filter(del_status=False).order_by('produto')
    numero_produtos = produtos.count()
    produtos = produtos[:100]

    conteudo = {
        'produtos': produtos,
        'TIPO_PRODUTO': TIPO_PRODUTO,
        'numero_produtos': numero_produtos,
    }
    return render(request, 'produtos/produtos.html', conteudo)

@login_required
def produtos_ficha(request, product_id=None):

    tags_selecionadas = []
    if product_id:
         try:
             produto = ProdutosFarmaceuticos.objects.get(id=product_id)
             tags_selecionadas = list(produto.produto_tag.active().values('tag_id', 'tag'))
         except ProdutosFarmaceuticos.DoesNotExist:
             messages.error(request, "Produto não encontrado.")
             return redirect('produtos')
    else:
         produto = None
    
    if request.method == 'POST':
        if produto:
            produto_form = ProdutosFarmaceuticosForm(request.POST, instance=produto)
        else:
            produto_form = ProdutosFarmaceuticosForm(request.POST)
        
        #Conferir se produto e tipo de produto foram preenchidos
        nome_produto = request.POST.get('produto')
        tipo_produto = request.POST.get('tipo_produto')
        if not nome_produto or not tipo_produto:
            messages.error(request, "O nome do produto é obrigatório!")
            if produto:
                return JsonResponse({
                    'redirect_url': reverse('produtos_ficha', args=[product_id]),
                })
            else:
                return JsonResponse({
                    'redirect_url': reverse('novo_produto'),
                })
        
        #Verificar se houve alteração no formulário
        if not produto_form.has_changed():
            messages.error(request, "Dados não foram salvos. Não houve mudanças.")
            if produto:
                return JsonResponse({
                    'redirect_url': reverse('produtos_ficha', args=[product_id]),
                })
            else:
                return JsonResponse({
                    'redirect_url': reverse('novo_produto'),
                })
        
        if produto_form.is_valid():
            #Verificar se já existe a produto na base
            nome_produto = produto_form.cleaned_data.get('produto')
            produto_existente = ProdutosFarmaceuticos.objects.filter(produto=nome_produto)
            
            #Se estivermos atualizando um produto existente, excluímos essa denominação da verificação
            if produto:
                produto_existente = produto_existente.exclude(id=produto.id)

            if produto_existente.exists():
                messages.error(request, "Já existe um produto com esse nome. Não foi possível salvar.")
                if produto:
                    return JsonResponse({
                        'redirect_url': reverse('produtos_ficha', args=[product_id]),
                    })
                else:
                    return JsonResponse({
                        'redirect_url': reverse('novo_produto'),
                    })

            #Salvar o produto
            produto = produto_form.save(commit=False)
            produto.save(current_user=request.user.usuario_relacionado)
            messages.success(request, "Dados atualizados com sucesso!")
            
            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Produtos Farmacêuticos_Produtos",
                model='ProdutosFarmaceuticos',
                model_id={produto.id},
                item_id=0,
                item_descricao="Salvar edição produto farmacêutico.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o produto farmacêutico {produto.produto} (ID: {produto.id}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar log
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'registro_data': produto.registro_data.strftime('%d/%m/%Y %H:%M:%S'),
                    'usuario_registro': produto.usuario_registro.dp_nome_completo,
                    'ult_atual_data': produto.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S'),
                    'usuario_atualizacao': produto.usuario_atualizacao.dp_nome_completo,
                    'log_n_edicoes': produto.log_n_edicoes,
                    'redirect_url': reverse('produtos_ficha', args=[product_id]),
                })
   
        else:
            messages.error(request, "Formulário inválido")
            print("Erro formulário produto")
            print(produto_form.errors)

    denominacoes_genericas = DenominacoesGenericas.objects.values_list('id', 'denominacao', 'tipo_produto')
    lista_atc = list(ListaATC.objects.values_list('codigo', 'descricao'))
    lista_atc.insert(0, ('Não informado', 'Não informado'))    
    tags_produtos = json.dumps(list(Tags.objects.values_list('id','tag')))
    form = ProdutosFarmaceuticosForm(instance=produto)
    
    return render(request, 'produtos/produtos_ficha.html', {
        'product_id': product_id,
        'produto': produto,
        'form': form,
        'denominacoes_genericas': denominacoes_genericas,
        'lista_atc': lista_atc,
        'tags_produtos': tags_produtos,
        'tags_selecionadas': tags_selecionadas,
        'TIPO_PRODUTO': TIPO_PRODUTO,
        'FORMA_FARMACEUTICA': FORMA_FARMACEUTICA, 
        'STATUS_INCORPORACAO': STATUS_INCORPORACAO,
        'CONCENTRACAO_TIPO': CONCENTRACAO_TIPO,
        'YES_NO': YES_NO,
        'CLASSIFICACAO_AWARE': CLASSIFICACAO_AWARE,
    })

def produtos_cmm(request, product_id=None):
    tab_cmm = ProdutoConsumoMedio.objects.filter(produto=product_id)
    conteudo = {
        'tab_cmm': tab_cmm,
    }
    
    return render(request, 'produtos/produtos_cmm.html', conteudo)

@login_required
def delete_produto(request, product_id):
    try:
        produto = ProdutosFarmaceuticos.objects.get(id=product_id)
        produto.soft_delete(request.user.usuario_relacionado)
        messages.error(request, "Produto deletado com sucesso.")

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Produtos Farmacêuticos",
            model='ProdutosFarmaceuticos',
            model_id={produto.id},
            item_id=0,
            item_descricao="Deleção de produto farmacêutico.",
            acao="Deleção",
            observacoes=f"Usuário {request.user.username} deletou o produto farmacêutico {produto.produto} (ID: {produto.id}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({"message": "Produto deletado com sucesso!"})
    except ProdutosFarmaceuticos.DoesNotExist:
        messages.error(request, "Produto não encontrado.")    
    return redirect('produtos')

@login_required
def get_filtros_produtos(request):
    tipo_produto = request.GET.get('tipo_produto', None)
    produto = request.GET.get('produto', None)
    basico = request.GET.get('basico', None)
    especializado = request.GET.get('especializado', None)
    estrategico = request.GET.get('estrategico', None)
    farmacia_popular = request.GET.get('farmacia_popular', None)
    hospitalar = request.GET.get('hospitalar', None)
    print('Basico: ', basico)
    filters = {}
    filters['del_status'] = False
    if tipo_produto:
        filters['denominacao__tipo_produto'] = tipo_produto
    if produto:
        filters['produto__icontains'] = produto
    if basico:
        filters['comp_basico'] = 1
    if especializado:
        filters['comp_especializado'] = 1
    if estrategico:
        filters['comp_estrategico'] = 1
    if farmacia_popular:
        filters['disp_farmacia_popular'] = 1
    if hospitalar:
        filters['hospitalar'] = 1
    
    produtos = ProdutosFarmaceuticos.objects.filter(**filters).order_by('produto')
    numero_produtos = produtos.count()
    
    page = int(request.GET.get('page', 1))
    paginator = Paginator(produtos, 100)  # Mostra 100 produtos por página
    try:
        produtos_paginados = paginator.page(page)
    except EmptyPage:
        produtos_paginados = paginator.page(paginator.num_pages)

    #data = list(produtos_paginados.object_list.values())
    data = list(produtos_paginados.object_list.values('id', 'denominacao__tipo_produto', 'produto', 'comp_basico', 'comp_especializado', 'comp_estrategico', 'disp_farmacia_popular', 'hospitalar'))

    return JsonResponse({
        'data': data,
        'numero_produtos': numero_produtos,
        'has_next': produtos_paginados.has_next(),
        'has_previous': produtos_paginados.has_previous(),
        'current_page': page
    })


@login_required
def exportar_produtos(request):
    print("Exportar Produtos")
    
    if request.method == 'POST':
        data = json.loads(request.body)
        tipo_produto = data.get('tipo_produto')
        produto = data.get('produto')
        basico = data.get('basico')
        especializado = data.get('especializado')
        estrategico = data.get('estrategico')
        farmacia_popular = data.get('farmacia_popular')
        hospitalar = data.get('hospitalar')
    
        filters = {}
        filters['del_status'] = False
        if tipo_produto:
            filters['denominacao__tipo_produto'] = tipo_produto
        if produto:
            filters['produto__icontains'] = produto
        if basico:
            filters['unidade_basico'] = basico
        if especializado:
            filters['unidade_especializado'] = especializado
        if estrategico:
            filters['unidade_estrategico'] = estrategico
        if farmacia_popular:
            filters['unidade_farm_popular'] = farmacia_popular
        if hospitalar:
            filters['hospitalar'] = hospitalar
        
        produtos = ProdutosFarmaceuticos.objects.filter(**filters)
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "denominacoes_genericas"

        headers = [
        'ID', 'Usuário Registro', 'Usuário Atualização', 'Data de Registro', 'Data da Última Atualização',
        'N Edições', 'Tipo Produto', 'Produto Farmacêutico', 'Concentração', 'Forma Farmacêutica', 'Oncológico', 'Biológico',
        'AWaRe', 'ATC - Código', 'ATC - Descrição', 'Incorporação SUS', 'Data Incorporação', 'Portaria Incorp.',
        'Link Portaria Incorp.', 'Data Exclusão', 'Portaria Exclusão', 'Link Exclusão',
        'Comp. Básico', 'Comp. Especializado', 'Comp. Estratégico', 'Disp. Farmácia Popular', 'Hospitalar',
        'SIGTAP', 'SIGTAP Cód.', 'SIGTAP Nome', 'SISMAT', 'SISMAT Cód.', 'SISMAT Nome', 'CATMAT', 'CATMAT Cód.', 'CATMAT Nome',
        'OBM', 'OBM Cód.', 'OBM Nome', 'Observações gerais', 'Data Exportação'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 15

        # Adicionar dados da tabela
        for row_num, produto in enumerate(produtos, 2):
            ws.cell(row=row_num, column=1, value=produto.id)
            ws.cell(row=row_num, column=2, value=str(produto.usuario_registro.primeiro_ultimo_nome()))
            ws.cell(row=row_num, column=3, value=str(produto.usuario_atualizacao.primeiro_ultimo_nome()))
            registro_data = produto.registro_data.replace(tzinfo=None)
            ult_atual_data = produto.ult_atual_data.replace(tzinfo=None)
            ws.cell(row=row_num, column=4, value=registro_data)
            ws.cell(row=row_num, column=5, value=ult_atual_data)
            ws.cell(row=row_num, column=6, value=produto.log_n_edicoes)
            ws.cell(row=row_num, column=7, value=produto.denominacao.tipo_produto)
            ws.cell(row=row_num, column=8, value=produto.produto)
            ws.cell(row=row_num, column=9, value=produto.concentracao)
            ws.cell(row=row_num, column=10, value=produto.get_forma_farmaceutica_display())
            ws.cell(row=row_num, column=11, value=produto.oncologico)
            ws.cell(row=row_num, column=12, value=produto.biologico)
            ws.cell(row=row_num, column=13, value=produto.aware)
            ws.cell(row=row_num, column=14, value=produto.atc)
            ws.cell(row=row_num, column=15, value=produto.atc_descricao)
            ws.cell(row=row_num, column=16, value=produto.incorp_status)
            ws.cell(row=row_num, column=17, value=produto.incorp_portaria)
            ws.cell(row=row_num, column=18, value=produto.incorp_link)
            ws.cell(row=row_num, column=19, value=produto.exclusao_data)
            ws.cell(row=row_num, column=20, value=produto.exclusao_portaria)
            ws.cell(row=row_num, column=21, value=produto.exclusao_link)
            ws.cell(row=row_num, column=22, value=produto.comp_basico)
            ws.cell(row=row_num, column=23, value=produto.comp_especializado)
            ws.cell(row=row_num, column=24, value=produto.comp_estrategico)
            ws.cell(row=row_num, column=25, value=produto.disp_farmacia_popular)
            ws.cell(row=row_num, column=26, value=produto.hospitalar)
            ws.cell(row=row_num, column=27, value=produto.sigtap_possui)
            ws.cell(row=row_num, column=28, value=produto.sigtap_codigo)
            ws.cell(row=row_num, column=29, value=produto.sigtap_nome)
            ws.cell(row=row_num, column=30, value=produto.sismat_possui)
            ws.cell(row=row_num, column=31, value=produto.sismat_codigo)
            ws.cell(row=row_num, column=32, value=produto.sismat_nome)
            ws.cell(row=row_num, column=33, value=produto.catmat_possui)
            ws.cell(row=row_num, column=34, value=produto.catmat_codigo)
            ws.cell(row=row_num, column=35, value=produto.catmat_nome)
            ws.cell(row=row_num, column=36, value=produto.obm_possui)
            ws.cell(row=row_num, column=37, value=produto.obm_codigo)
            ws.cell(row=row_num, column=38, value=produto.obm_nome)
            ws.cell(row=row_num, column=39, value=produto.observacoes_gerais)
            ws.cell(row=row_num, column=15, value=current_date_str)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Reposition to the start of the stream

        # Registrar a ação no CustomLog
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Produtos Farmacêuticos_Produtos",
            model='ProdutosFarmaceuticos',
            model_id=0,
            item_id=0,
            item_descricao="Exportação de produtos farmacêuticos",
            acao="Exportação",
            observacoes=f"Usuário {request.user.username} exportou produtos farmacêuticos em {current_date_str}."
        )
        log_entry.save()

        # Configurar a resposta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportar_denominacoes.xlsx"'
        response.write(output.getvalue())
        return response

@login_required
def salvar_tags(request, product_id):
    if request.method == "POST":
        produto = get_object_or_404(ProdutosFarmaceuticos, pk=product_id)

        # Decodificar as tags enviadas de uma string JSON para uma lista de dicionários
        tags_selecionadas_json = request.POST.get('tags')
        tags_selecionadas_list = json.loads(tags_selecionadas_json)

        # Obter apenas os IDs das tags selecionadas para comparação
        tags_selecionadas_ids = [tag["id"] for tag in tags_selecionadas_list]

        # Tags atuais do produto
        tags_banco_ids = list(produto.produto_tag.active().values_list('tag_id', flat=True))
        
        # Tags para salvar
        tags_salvar = [tag for tag in tags_selecionadas_list if tag["id"] not in tags_banco_ids]

        # Usuario
        usuario_instance = request.user.usuario_relacionado
        
        # Tags para adicionar
        contador = 0
        print("Contador = ", contador)
        for tag_data in tags_salvar:
            tag_instance = ProdutosTags(
                usuario_registro=usuario_instance,
                usuario_atualizacao=usuario_instance,
                produto=produto,
                tag=tag_data["value"],
                tag_id=tag_data["id"]
            )
            contador = contador + 1
            print("Contador = ", contador)
            tag_instance.save()

        # Tags para deletar (soft delete)
        tags_deletar = [tag_id for tag_id in tags_banco_ids if tag_id not in tags_selecionadas_ids]
        
        for tag_id in tags_deletar:
            tag_instance = ProdutosTags.objects.filter(produto=produto, tag_id=tag_id, del_status=False).first()
            print("Chamando soft_delete para tag_id:", tag_id)
            if tag_instance:
                tag_instance.soft_delete(usuario_instance)

        # Registrar a ação no CustomLog
        tags_nomes_para_log = [tag['value'] for tag in tags_selecionadas_list]
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Produtos Farmacêuticos_Produtos_Tags",
            model='ProdutosTags',
            model_id={tags_nomes_para_log},
            item_id=0,
            item_descricao="Salvar edição de tags de produto farmacêutico.",
            acao="Salvar",
            observacoes=f"Usuário {request.user.username} salvou tags ({', '.join(tags_nomes_para_log)}) do produto farmacêutico {produto.produto} (ID: {produto.id}) em {current_date_str}."
        )
        log_entry.save()

        messages.success(request, f"Tags atualizadas com sucesso!")
        return JsonResponse({"status": "success"})
    else:
        messages.success(request, f"Erro! Tags não atualizadas!")
        return JsonResponse({"status": "error", "message": "Invalid request method"})

@login_required
def produto_exportar_pdf(request, product_id):
    pass
#     #Buscar o produto
#     produto = ProdutosFarmaceuticos.objects.get(id=product_id)

#     # Renderizar o template com os produtos
#     html_string = render_to_string('templates/_outputs/output_ficha_produto.html', {'produto': produto})

#     # Gerar o PDF
#     # response = HttpResponse(content_type='application/pdf')
#     # response['Content-Disposition'] = 'attachment; filename="produtos.pdf"'
#     # html = HTML(string=html_string)
#     # pisa_status = pisa.CreatePDF(html_string, dest=response)

#     # if pisa_status.err:
#     #     return HttpResponse('Tivemos alguns erros ao gerar o PDF <pre>' + html_string + '</pre>')
#     # return response




@login_required
def denominacoes(request):
    denominacoes = DenominacoesGenericas.objects.filter(del_status=False).order_by('denominacao')
    numero_denominacoes = denominacoes.count()
    context = {
        'denominacoes': denominacoes,
        'TIPO_PRODUTO': TIPO_PRODUTO,
        'numero_denominacoes': numero_denominacoes,
    }
    return render(request, 'produtos/denominacoes.html', context)

@login_required
def denominacoes_ficha(request, denominacao_id=None):

    if denominacao_id:
        try:
            denominacao = DenominacoesGenericas.objects.get(id=denominacao_id)
        except DenominacoesGenericas.DoesNotExist:
            messages.error(request, "Denominação não encontrada.")
            return redirect('denominacoes')
    else:
        denominacao = None  # Preparando para criar uma nova denominação

    if request.method == 'POST':
        if denominacao:
            denominacao_form = DenominacoesGenericasForm(request.POST, instance=denominacao)
        else:
            denominacao_form = DenominacoesGenericasForm(request.POST)

        #Conferir se denominação e tipo de produto foram preenchidos
        nome_denominacao = request.POST.get('denominacao')
        tipo_produto = request.POST.get('tipo_produto')
        if not nome_denominacao or not tipo_produto:
            messages.error(request, "O nome da denominação genérica e o tipo de produto são obrigatórios!")
            if denominacao:
                return redirect('denominacoes_ficha', denominacao_id=denominacao.id)
            else:
                return redirect('nova_denominacao')

        #Verificar se houve alteração no formulário
        if not denominacao_form.has_changed():
            messages.error(request, "Dados não foram salvos. Não houve mudanças.")
            if denominacao:
                return redirect('denominacoes_ficha', denominacao_id=denominacao.id)
            else:
                return redirect('nova_denominacao')

        #Formulário é válido
        if denominacao_form.is_valid():
            #Verificar se já existe a denominação na base [
            nome_denominacao = denominacao_form.cleaned_data.get('denominacao')
            denominacao_existente = DenominacoesGenericas.objects.filter(denominacao=nome_denominacao)
            
            #Se estivermos atualizando uma denominação existente, excluímos essa denominação da verificação
            if denominacao:
                denominacao_existente = denominacao_existente.exclude(id=denominacao.id)

            if denominacao_existente.exists():
                messages.error(request, "Já existe uma denominação com esse nome. Não foi possível salvar.")
                if denominacao:
                    return redirect('denominacoes_ficha', denominacao_id=denominacao.id)
                else:
                    return redirect('nova_denominacao')

            #Salvar a denominação
            denominacao = denominacao_form.save(commit=False)
            denominacao.save(current_user=request.user.usuario_relacionado)
            messages.success(request, f"Dados atualizados com sucesso!")
            
            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Produtos Farmacêuticos_Denominações Genéricas",
                model='DenominacoesGenericas',
                model_id={denominacao.id},
                item_id=0,
                item_descricao="Salvar edição denominação genérica.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou a denominação genéria {denominacao.denominacao} (ID: {denominacao.id}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar log
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'registro_data': denominacao.registro_data.strftime('%d/%m/%Y %H:%M:%S'),
                    'usuario_registro': denominacao.usuario_registro.dp_nome_completo,
                    'ult_atual_data': denominacao.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S'),
                    'usuario_atualizacao': denominacao.usuario_atualizacao.dp_nome_completo,
                    'log_n_edicoes': denominacao.log_n_edicoes,
                    'id': denominacao.id,
                })           
            
        else:
            messages.error(request, "Formulário inválido")
            print("Erro formulário denominação")
            print(denominacao_form.errors)

    form = DenominacoesGenericasForm(instance=denominacao)
    return render(request, 'produtos/denominacoes_ficha.html', {
        'denominacao': denominacao,
        'form': form,
        'TIPO_PRODUTO': TIPO_PRODUTO,
    })

@login_required
def delete_denominacao(request, denominacao_id):
    try:
        denominacao = DenominacoesGenericas.objects.get(id=denominacao_id)
        denominacao.soft_delete(request.user.usuario_relacionado)
        messages.error(request, "Denominação deletada com sucesso!")

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Produtos Farmacêuticos_Denominações Genéricas",
            model='DenominacoesGenericas',
            model_id={denominacao.id},
            item_id=0,
            item_descricao="Deleção de denominação genérica.",
            acao="Deleção",
            observacoes=f"Usuário {request.user.username} deletou a denominação genéria {denominacao.denominacao} (ID: {denominacao.id}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({"message": "Denominação deletada com sucesso!"})
    except DenominacoesGenericas.DoesNotExist:
        messages.error(request, "Denominação não encontrada.")    
    return redirect('denominacoes')

@login_required
def get_filtros_denominacoes(request):
    tipo_produto = request.GET.get('tipo_produto', None)
    denominacao = request.GET.get('denominacao', None)
    basico = request.GET.get('basico', None)
    especializado = request.GET.get('especializado', None)
    estrategico = request.GET.get('estrategico', None)
    farmacia_popular = request.GET.get('farmacia_popular', None)
    hospitalar = request.GET.get('hospitalar', None)

    filters = {}
    filters['del_status'] = False
    if tipo_produto:
        filters['tipo_produto'] = tipo_produto
    if denominacao:
        filters['denominacao__icontains'] = denominacao
    if basico:
        filters['unidade_basico'] = 1
    if especializado:
        filters['unidade_especializado'] = 1
    if estrategico:
        filters['unidade_estrategico'] = 1
    if farmacia_popular:
        filters['unidade_farm_popular'] = 1
    if hospitalar:
        filters['hospitalar'] = 1
    
    denominacoes = DenominacoesGenericas.objects.filter(**filters).order_by('denominacao')
    numero_denominacoes = denominacoes.count()
    
    page = int(request.GET.get('page', 1))
    paginator = Paginator(denominacoes, 100)  # Mostra 100 denominações por página
    try:
        denominacoes_paginados = paginator.page(page)
    except EmptyPage:
        denominacoes_paginados = paginator.page(paginator.num_pages)

    data = list(denominacoes_paginados.object_list.values())

    return JsonResponse({
        'data': data,
        'numero_denominacoes': numero_denominacoes,
        'has_next': denominacoes_paginados.has_next(),
        'has_previous': denominacoes_paginados.has_previous(),
        'current_page': page
    })

@login_required
def exportar_denominacoes(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        tipo_produto = data.get('tipo_produto')
        denominacao = data.get('denominacao')
        basico = data.get('basico')
        especializado = data.get('especializado')
        estrategico = data.get('estrategico')
        farmacia_popular = data.get('farmacia_popular')
        hospitalar = data.get('hospitalar')
    
        filters = {}
        filters['del_status'] = False
        if tipo_produto:
            filters['tipo_produto'] = tipo_produto
        if denominacao:
            filters['denominacao__icontains'] = denominacao
        if basico:
            filters['unidade_basico'] = basico
        if especializado:
            filters['unidade_especializado'] = especializado
        if estrategico:
            filters['unidade_estrategico'] = estrategico
        if farmacia_popular:
            filters['unidade_farm_popular'] = farmacia_popular
        if hospitalar:
            filters['hospitalar'] = hospitalar
        

        denominacoes = DenominacoesGenericas.objects.filter(**filters)
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "denominacoes_genericas"

        headers = [
        'ID', 'Usuário Registro', 'Usuário Atualização', 'Data de Registro', 'Data da Última Atualização',
        'Denominação', 'Tipo de Produto', 'Unidade Básico', 'Unidade Especializado', 'Unidade Estratégico',
        'Unidade Farmácia Popular', 'Hospitalar', 'Observações Gerais', 'Data Exportação'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 15

        # Adicionar dados da tabela
        for row_num, denominacao in enumerate(denominacoes, 2):
            ws.cell(row=row_num, column=1, value=denominacao.id)
            ws.cell(row=row_num, column=2, value=str(denominacao.usuario_registro.primeiro_ultimo_nome()))
            ws.cell(row=row_num, column=3, value=str(denominacao.usuario_atualizacao.primeiro_ultimo_nome()))
            registro_data = denominacao.registro_data.replace(tzinfo=None)
            ult_atual_data = denominacao.ult_atual_data.replace(tzinfo=None)
            ws.cell(row=row_num, column=4, value=registro_data)
            ws.cell(row=row_num, column=5, value=ult_atual_data)
            ws.cell(row=row_num, column=6, value=denominacao.denominacao)
            ws.cell(row=row_num, column=7, value=denominacao.tipo_produto)
            ws.cell(row=row_num, column=8, value=denominacao.unidade_basico)
            ws.cell(row=row_num, column=9, value=denominacao.unidade_especializado)
            ws.cell(row=row_num, column=10, value=denominacao.unidade_estrategico)
            ws.cell(row=row_num, column=11, value=denominacao.unidade_farm_popular)
            ws.cell(row=row_num, column=12, value=denominacao.hospitalar)
            ws.cell(row=row_num, column=13, value=denominacao.observacoes_gerais)
            ws.cell(row=row_num, column=14, value=current_date_str)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Reposition to the start of the stream

        # Registrar a ação no CustomLog
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Denominações Genéricas",
            model='DenominacoesGenericas',
            model_id=0,
            item_id=0,
            item_descricao="Exportação de denominações genéricas",
            acao="Exportação",
            observacoes=f"Usuário {request.user.username} exportou denominações genéricas em {current_date_str}."
        )
        log_entry.save()

        # Configurar a resposta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportar_denominacoes.xlsx"'
        response.write(output.getvalue())
        return response

@login_required
def denominacoes_buscar(request, unidade_daf=None):
    if unidade_daf:
        print('Buscando denominações para a unidade:', unidade_daf)
        denominacoes = [d for d in DenominacoesGenericas.objects.all() if unidade_daf in d.componentes_af]
    else:
        denominacoes = DenominacoesGenericas.objects.filter(del_status=False).order_by('denominacao')

    # Construindo a lista de dicionários manualmente
    denominacoes_list = [{'id': d.id, 'denominacao': d.denominacao} for d in denominacoes]

    return JsonResponse({'denominacoes_list': denominacoes_list})





