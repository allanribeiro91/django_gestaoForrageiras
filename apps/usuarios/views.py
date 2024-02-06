from django import forms
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from setup.choices import GENERO_SEXUAL, COR_PELE, VINCULO_MS, ORGAO_PUBLICO
from datetime import datetime
from apps.usuarios.forms import UsuarioForms
from apps.main.models import CustomLog, UserAccessLog
import pdb
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import json

@login_required
def meusdados(request):
    usuario = request.user.usuario_relacionado
    alocacao_ativa = usuario.alocacao_ativa()
    
    if request.method == 'POST':
        if editar_meusdados(request):
            return redirect('meusdados')
    
    form = UsuarioForms(instance=usuario)
    return render(request, 'usuarios/meusdados.html', {
        'usuario': usuario,
        'form': form,
        'alocacao_ativa': alocacao_ativa,
        'GENERO_SEXUAL': GENERO_SEXUAL,
        'COR_PELE': COR_PELE,
    })

@login_required
def meuslogs(request):
    usuario = request.user.usuario_relacionado
    tab_logs = CustomLog.objects.filter(usuario=usuario).order_by('-timestamp')
    return render(request, 'usuarios/meuslogs.html', {
        'usuario': usuario,
        'tab_logs': tab_logs,
    })

def meuslogs_exportar(request):
    print("Exportar Logs")
    
    if request.method == 'POST':
        usuario = request.user.usuario_relacionado
        tab_logs = CustomLog.objects.filter(usuario=usuario).order_by('-timestamp')
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "logs_sisdaf"

        headers = [
        'ID', 'Usuário', 'CPF', 'Data do Log', 'Módulo SisDAF', 'Ação', 'Descrição da ação', 'Detalhamento', 'Data Exportação'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 20

        # Adicionar dados da tabela
        for row_num, log in enumerate(tab_logs, 2):
            ws.cell(row=row_num, column=1, value=log.id)
            ws.cell(row=row_num, column=2, value=log.usuario.dp_nome_completo)
            ws.cell(row=row_num, column=3, value=log.usuario.dp_cpf)
            data_log = log.timestamp.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M:%S')
            ws.cell(row=row_num, column=4, value=data_log)
            ws.cell(row=row_num, column=5, value=log.modulo)
            ws.cell(row=row_num, column=6, value=log.acao)
            ws.cell(row=row_num, column=7, value=log.item_descricao)
            ws.cell(row=row_num, column=8, value=log.observacoes)
            ws.cell(row=row_num, column=9, value=current_date_str)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Registrar a ação no CustomLog
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Usuário",
            model='CustomLog',
            model_id=0,
            item_id=0,
            item_descricao="Exportação da lista de logs (ações) do usuário no SisDAF.",
            acao="Exportação",
            observacoes=f"Usuário {request.user.username} exportou lista dos seus logs (ações) no SisDAF em {current_date_str}."
        )
        log_entry.save()

        # Configurar a resposta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportar_logs_sisdaf.xlsx"'
        response.write(output.getvalue())
        return response


def meusacessos(request):
    usuario = request.user.usuario_relacionado
    tab_acessos = UserAccessLog.objects.filter(usuario=usuario).order_by('-timestamp')
    return render(request, 'usuarios/meusacessos.html', {
        'usuario': usuario,
        'tab_acessos': tab_acessos,
    })

def meusacessos_exportar(request):
    print("Exportar Acessos")
    
    if request.method == 'POST':
        usuario = request.user.usuario_relacionado
        tab_acessos = UserAccessLog.objects.filter(usuario=usuario).order_by('-timestamp')
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "acessos_sisdaf"

        headers = [
        'ID', 'Data do acesso', 'Usuário', 'CPF', 'Data Exportação'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 25

        # Adicionar dados da tabela
        for row_num, acesso in enumerate(tab_acessos, 2):
            ws.cell(row=row_num, column=1, value=acesso.id)
            data_acesso = acesso.timestamp.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M:%S')
            ws.cell(row=row_num, column=2, value=data_acesso)
            ws.cell(row=row_num, column=3, value=acesso.usuario.dp_nome_completo)
            ws.cell(row=row_num, column=4, value=acesso.usuario.dp_cpf)
            ws.cell(row=row_num, column=5, value=current_date_str)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Reposition to the start of the stream

        # Registrar a ação no CustomLog
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Usuário",
            model='UserAccessLog',
            model_id=0,
            item_id=0,
            item_descricao="Exportação da lista de acessos do usuário ao SisDAF.",
            acao="Exportação",
            observacoes=f"Usuário {request.user.username} exportou lista dos seus acessos ao SisDAF em {current_date_str}."
        )
        log_entry.save()

        # Configurar a resposta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportar_acessos_sisdaf.xlsx"'
        response.write(output.getvalue())
        return response

@login_required
def editar_meusdados(request):
    usuario = request.user.usuario_relacionado
    print(usuario)
    if request.method == 'POST':
        usuario_form = UsuarioForms(request.POST, request.FILES, instance=usuario)

        if usuario_form.is_valid():
            usuario_form.save()
            messages.success(request, f"Dados atualizados com sucesso!")
            return True

        else:
            messages.error(request, "Formulário inválido")
            print(usuario_form.errors)
            return False

