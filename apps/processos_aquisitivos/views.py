from unidecode import unidecode
from django.core.paginator import Paginator, EmptyPage
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.template.loader import render_to_string
from django.contrib import auth, messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from setup.choices import UNIDADE_DAF2, MODALIDADE_AQUISICAO, STATUS_PROAQ, STATUS_FASE
from apps.usuarios.models import Usuarios
from apps.produtos.models import DenominacoesGenericas, ProdutosFarmaceuticos
from apps.main.models import CustomLog
from apps.processos_aquisitivos.models import ProaqDadosGerais, ProaqProdutos, ProaqEvolucao, PROAQ_AREA_MS, PROAQ_ETAPA, ProaqTramitacao
from apps.processos_aquisitivos.forms import ProaqDadosGeraisForm, ProaqEvolucaoForm, ProaqTramitacaoForm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
import re
from datetime import datetime

def proaq(request):
    tab_proaqs = ProaqDadosGerais.objects.filter(del_status=False).order_by('denominacao')
    lista_denominacoes = (
        ProaqDadosGerais.objects.filter(del_status=False)
        .values_list('denominacao__id', 'denominacao__denominacao')
        .distinct()
    )
    total_processos = tab_proaqs.count()
    # Ordenando a lista em Python, removendo acentos e considerando maiúsculas/minúsculas
    lista_denominacoes = sorted(lista_denominacoes, key=lambda x: unidecode((x[1] or '').lower()))
    conteudo = {
        'tab_proaqs': tab_proaqs,
        'MODALIDADE_AQUISICAO': MODALIDADE_AQUISICAO,
        'STATUS_PROAQ': STATUS_PROAQ,
        'UNIDADE_DAF': UNIDADE_DAF2,
        'DENOMINACOES': lista_denominacoes,
        'total_processos': total_processos,
    }
    return render(request, 'processos_aquisitivos/proaq.html', conteudo)

def proaq_ficha_dados_gerais(request, proaq_id=None):
    
    produtos_selecionados = []
    nomes_produtos = None
    if proaq_id:
         try:
             proaq_dados_gerais = ProaqDadosGerais.objects.get(id=proaq_id)
             produtos_selecionados = list(proaq_dados_gerais.proaq_produto.filter(del_status=False).values('produto_id', 'produto__produto'))
             nomes_produtos = [produto['produto__produto'] for produto in produtos_selecionados]
         except ProaqDadosGerais.DoesNotExist:
             messages.error(request, "Processo Aquisitivo não encontrado.")
             return redirect('proaq')
    else:
         proaq_dados_gerais = None  # Preparando para criar um novo processo

    if request.method == 'POST':

        if proaq_dados_gerais:
            proaq_dados_gerais_form = ProaqDadosGeraisForm(request.POST, instance=proaq_dados_gerais)
            novo_proaq = False
        else:
            proaq_dados_gerais_form = ProaqDadosGeraisForm(request.POST)
            novo_proaq = True

        #Verificar se houve alteração no formulário
        if not proaq_dados_gerais_form.has_changed():
            messages.error(request, "Dados não foram salvos. Não houve mudanças.")
            if proaq_dados_gerais:
                return redirect('proaq_ficha_dados_gerais', proaq_id=proaq_dados_gerais.id)
            else:
                return redirect('novo_proaq')

        if proaq_dados_gerais_form.is_valid():
            #Verificar se já existe registro desse processo aquisitivo
            numero_processo_sei = proaq_dados_gerais_form.cleaned_data.get('numero_processo_sei')
            processo_sei_existente = ProaqDadosGerais.objects.filter(numero_processo_sei=numero_processo_sei)

            #Se estivermos atualizando um processo existente, excluímos esse processo da verificação
            if proaq_dados_gerais:
                processo_sei_existente = processo_sei_existente.exclude(id=proaq_dados_gerais.id)
            
            if processo_sei_existente.exists():
                messages.error(request, "Já existe um registro com esse Processo SEI. Não foi possível salvar.")
                if proaq_dados_gerais:
                    return redirect('proaq_ficha_dados_gerais', proaq_id=proaq_dados_gerais.id)
                else:
                    return redirect('novo_proaq')

            #Salvar o produto
            proaq_dados_gerais = proaq_dados_gerais_form.save(commit=False)
            proaq_dados_gerais.save(current_user=request.user.usuario_relacionado)
            
            if novo_proaq:
                messages.success(request, "Novo processo aquisitivo registrado com sucesso!")
            else:
                messages.success(request, "Dados atualizados com sucesso!")
                print("Dado atualizado")

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Processo Aquisitivo_Dados Gerais",
                model='ProaqDadosGerais',
                model_id={proaq_dados_gerais.id},
                item_id=0,
                item_descricao="Salvar edição de processo aquisitivo.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o processo aquisitivo (ID: {proaq_dados_gerais.id}) da unidade daf {proaq_dados_gerais.unidade_daf} e denominação genérica {proaq_dados_gerais.denominacao.denominacao} em {current_date_str}."
            )
            log_entry.save()

            #Retornar log
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'redirect_url': reverse('proaq_ficha_dados_gerais', args=[proaq_dados_gerais.id]),
                    'registro_data': proaq_dados_gerais.registro_data.strftime('%d/%m/%Y %H:%M:%S'),
                    'usuario_registro': proaq_dados_gerais.usuario_registro.dp_nome_completo,
                    'ult_atual_data': proaq_dados_gerais.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S'),
                    'usuario_atualizacao': proaq_dados_gerais.usuario_atualizacao.dp_nome_completo,
                    'log_n_edicoes': proaq_dados_gerais.log_n_edicoes,
                    'id': proaq_dados_gerais.id,
                })
        else:
            messages.error(request, "Preencha os campos obrigatórios.")
            print("Erro formulário produto")
            print(proaq_dados_gerais_form.errors)
            
    denominacoes_genericas = DenominacoesGenericas.objects.filter(del_status=False).values_list('id', 'denominacao')
    form = ProaqDadosGeraisForm(instance=proaq_dados_gerais)

    return render(request, 'processos_aquisitivos/proaq_ficha_dados_gerais.html', {
        'denominacoes_genericas': denominacoes_genericas,
        'form': form,
        'proaq_dados_gerais': proaq_dados_gerais,
        'produtos_selecionados': produtos_selecionados,
        'nomes_produtos': nomes_produtos,
        'UNIDADE_DAF': UNIDADE_DAF2,
        'MODALIDADE_AQUISICAO': MODALIDADE_AQUISICAO,
        'STATUS_PROAQ': STATUS_PROAQ,
    })

def proaq_usuarios_por_unidade(request, unidade):
    usuarios = Usuarios.usuarios_por_unidade(unidade)
    return JsonResponse(usuarios, safe=False)

def proaq_produtos_por_denominacao(request, denominacao):
    produtos = ProdutosFarmaceuticos.get_produtos_por_denominacao(denominacao)
    return JsonResponse(produtos, safe=False)

def proaq_ficha_evolucao(request, proaq_id=None):
    form = ProaqEvolucaoForm()

    if request.method == 'POST':
        # 1. Extrair o id_proaq e os dados das fases do request
        id_proaq = request.POST.get('id_proaq')
        fases_data = json.loads(request.POST.get('fasesData'))

        # 2. Para cada fase recebida, verificar se já existe na base de dados. Se existir, atualizar. Se não, criar uma nova.
        proaq = get_object_or_404(ProaqDadosGerais, pk=id_proaq)
        user = request.user
        try:
            usuario_atualizacao = Usuarios.objects.get(user=user)  # Convertendo de User para Usuario
        except Usuarios.DoesNotExist:
            usuario_atualizacao = None
        
        print("PROAQ: ", id_proaq)
        log_n_edicoes = ProaqEvolucao.objects.filter(del_status=False, proaq_id=id_proaq, fase=1).first()
        if log_n_edicoes:
            log_n = log_n_edicoes.log_n_edicoes + 1
        else:
            log_n = 1

        created = False
        fase = None

        for fase_num, fase_data in fases_data.items():
            # Extrair o número da fase
            num = int(re.search(r'\d+', fase_num).group())

            data_inicio = fase_data.get('dataInicio')
            data_fim = fase_data.get('dataFim')

            fase, created = ProaqEvolucao.objects.update_or_create(
                proaq=proaq,
                fase=num,
                del_status=False,
                defaults={
                    'status': fase_data['status'],
                    'data_inicio': data_inicio if data_inicio else None,
                    'data_fim': data_fim if data_fim else None,
                    'comentario': fase_data.get('comentario'),
                    'usuario_atualizacao': usuario_atualizacao,
                    'usuario_registro': usuario_atualizacao if created or fase is None else fase.usuario_registro,
                    'log_n_edicoes': log_n,
                }
            )

            if fase is None:
                # Houve um problema ao tentar criar ou atualizar o objeto ProaqEvolucao.
                # Você pode decidir como lidar com essa situação.
                pass

        # # 3. Verificar se existem fases na base de dados que não foram recebidas e, se sim, fazer o soft delete dessas fases.
        existing_fases = ProaqEvolucao.objects.filter(proaq=proaq, del_status=False)
        received_fases = set(fases_data.keys())
        for fase in existing_fases:
            fase_num_str = f'fase{fase.fase}'
            if fase_num_str not in received_fases:
                fase.soft_delete(usuario_atualizacao)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Processo Aquisitivo_Evolução",
            model='ProaqEvolucao',
            model_id={proaq.id},
            item_id=0,
            item_descricao="Salvou evolução de processo aquisitivo.",
            acao="Salvar",
            observacoes=f"Usuário {request.user.username} salvou evolução o processo aquisitivo (ID: {proaq.id}) da unidade daf {proaq.unidade_daf} e denominação genérica {proaq.denominacao.denominacao} em {current_date_str}."
        )
        log_entry.save()

        #Retornar
        messages.success(request, "Dados atualizados com sucesso!")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'redirect_url': reverse('proaq_ficha_evolucao', args=[id_proaq]),
            })

        pass

    fases = [
        {'nome': 'Configuração dos documentos iniciais'},
        {'nome': 'Tramitação Pré-Contratual'},
        {'nome': 'Contratação'},
        {'nome': 'Execução do Contrato'},
        {'nome': 'Aditivo de Contrato'},
        {'nome': 'Encerramento'},
        {'nome': 'Processo Finalizado'},
    ]

    dados_fases = list(ProaqEvolucao.objects.filter(del_status=False, proaq_id=proaq_id))
    dados_fase1 = ProaqEvolucao.objects.filter(del_status=False, proaq_id=proaq_id, fase=1).first()

    conteudo = {
        'status_fase': STATUS_FASE,
        'form': form,
        'fases': fases,
        'proaq_id': proaq_id,
        'dados_fases': dados_fases,
        'dados_fase1': dados_fase1,
    }

    return render(request, 'processos_aquisitivos/proaq_ficha_evolucao.html', conteudo)

@login_required
def proaq_ficha_tramitacoes(request, tramitacao_id=None, proaq_id=None):

    tramitacao_id = request.POST.get('id_tramitacao')
    print('ID da Tramitação: ', tramitacao_id)

    if tramitacao_id:
        try:
            proaq_tramitacao = ProaqTramitacao.objects.get(id=tramitacao_id)
        except ProaqTramitacao.DoesNotExist:
            messages.error(request, "Tramitação não encontrada.")
            return redirect('proaq')
    else:
        proaq_tramitacao = None
    
    if request.method == 'POST':
        if proaq_tramitacao:
            proaq_tramitacao_form = ProaqTramitacaoForm(request.POST, instance=proaq_tramitacao)
            nova_tramitacao = False
        else:
            proaq_tramitacao_form = ProaqTramitacaoForm(request.POST)
            nova_tramitacao = True
            
        #Verificar se houve alteração no formulário
        if not proaq_tramitacao_form.has_changed():
            messages.error(request, "Dados não foram salvos. Não houve mudanças.")
            return redirect('proaq_ficha_tramitacoes', proaq_id=proaq_id)

        proaq_instance = ProaqDadosGerais.objects.get(id=proaq_id)
        proaq_tramitacao_form = ProaqTramitacaoForm(request.POST, instance=proaq_tramitacao)
        proaq_tramitacao_form.instance.proaq = proaq_instance

        if proaq_tramitacao_form.is_valid():
            #Verificar se já existe registro dessa tramitacao
            documento_sei = proaq_tramitacao_form.cleaned_data.get('documento_sei')
            documento_sei_existente = ProaqTramitacao.objects.filter(documento_sei=documento_sei)

            #Se estivermos atualizando um processo existente, excluímos esse processo da verificação
            if proaq_tramitacao:
                documento_sei_existente = documento_sei_existente.exclude(id=proaq_tramitacao.id)
            
            if documento_sei_existente.exists():
                messages.error(request, "Já existe uma tramitação com esse Documento SEI. Não foi possível salvar.")
                return redirect('proaq_ficha_tramitacoes', proaq_id=proaq_id)
            
            #Salvar a tramitação
            proaq_tramitacao = proaq_tramitacao_form.save(commit=False)
            proaq_tramitacao.save(current_user=request.user.usuario_relacionado)
            
            if nova_tramitacao:
                messages.success(request, "Nova tramitação registrada com sucesso!")
            else:
                messages.success(request, "Dados atualizados com sucesso!")
                print("Dado atualizado")

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Processo Aquisitivo_Tramitação",
                model='ProaqTramitacao',
                model_id={proaq_tramitacao.id},
                item_id=0,
                item_descricao="Salvou tramitação de processo aquisitivo.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou tramitação (ID: {proaq_tramitacao.id}) do processo aquisitivo (ID: {proaq_tramitacao.proaq.id}) da unidade daf {proaq_tramitacao.proaq.unidade_daf} e denominação genérica {proaq_tramitacao.proaq.denominacao.denominacao} em {current_date_str}."
            )
            log_entry.save()

            #Retornar log
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'redirect_url': reverse('proaq_ficha_tramitacoes', args=[proaq_id]),
                })
        else:
            messages.error(request, "Preencha os campos obrigatórios.")
            print("Erro formulário tramitação.")
            print(proaq_tramitacao_form.errors)

    if proaq_id:
        tab_tramitacoes = ProaqTramitacao.objects.filter(del_status=False, proaq=proaq_id).order_by('-data_entrada')
        for tramitacao in tab_tramitacoes:
            if tramitacao.data_saida and tramitacao.data_entrada:
                tramitacao.duracao = (tramitacao.data_saida - tramitacao.data_entrada).days
            elif tramitacao.data_entrada:
                tramitacao.duracao = (datetime.today().date() - tramitacao.data_entrada).days
            else:
                tramitacao.duracao = None
    else:
        tab_tramitacoes = None

    lista_area_ms = PROAQ_AREA_MS.objects.values_list('setor', flat=True).distinct().order_by('setor')
    lista_etapa = PROAQ_ETAPA.objects.values_list('etapa', flat=True).distinct().order_by('etapa')    

    return render(request, 'processos_aquisitivos/proaq_ficha_tramitacoes.html', {
        'lista_area_ms': lista_area_ms,
        'lista_etapa': lista_etapa,
        'tab_tramitacoes': tab_tramitacoes,
        'proaq_id': proaq_id,
    })

@login_required
def tramitacao_dados(request, tramitacao_id):
    try:
        tramitacao = ProaqTramitacao.objects.get(id=tramitacao_id)
        data = {
            'id_tramitacao': tramitacao.id,
            'log_data_registro': tramitacao.registro_data.strftime('%d/%m/%Y %H:%M:%S') if tramitacao.registro_data else '',
            'log_responsavel_registro': str(tramitacao.usuario_atualizacao.dp_nome_completo),
            'lot_ult_atualizacao': tramitacao.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if tramitacao.ult_atual_data else '',
            'log_responsavel_atualizacao': str(tramitacao.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': tramitacao.log_n_edicoes,
            'documento_sei': tramitacao.documento_sei,
            'tramitacao_local': tramitacao.setor,
            'tramitacao_etapa': tramitacao.etapa_processo,
            'tramitacao_data_entrada': tramitacao.data_entrada.strftime('%Y-%m-%d') if tramitacao.data_entrada else '',
            'tramitacao_data_previsao': tramitacao.previsao_saida.strftime('%Y-%m-%d') if tramitacao.previsao_saida else '',
            'tramitacao_data_saida': tramitacao.data_saida.strftime('%Y-%m-%d') if tramitacao.data_saida else '',
            'tramitacao_observacoes': tramitacao.observacoes,
        }
        return JsonResponse(data)
    except ProaqTramitacao.DoesNotExist:
        return JsonResponse({'error': 'Tramitação não encontrada'}, status=404)

@login_required
def proaq_filtro(request):
    status = request.GET.get('status', None)
    modalidadeAquisicao = request.GET.get('modalidadeAquisicao', None)
    unidadeDAF = request.GET.get('unidadeDAF', None)
    denominacao_id = request.GET.get('denominacao_id', None)

    filters = {}
    filters['del_status'] = False
    if status:
        filters['status'] = status
    if modalidadeAquisicao:
        filters['modalidade_aquisicao'] = modalidadeAquisicao
    if unidadeDAF:
        filters['unidade_daf'] = unidadeDAF
    if denominacao_id:
        filters['denominacao_id'] = denominacao_id
    
    tab_proaqs = ProaqDadosGerais.objects.filter(**filters).order_by('denominacao')
    total_processos = tab_proaqs.count()
    
    page = int(request.GET.get('page', 1))
    paginator = Paginator(tab_proaqs, 100)  # Mostra 100 denominações por página
    try:
        proaqs_paginados = paginator.page(page)
    except EmptyPage:
        proaqs_paginados = paginator.page(paginator.num_pages)

    data = list(proaqs_paginados.object_list.values())

    # Adicionando os dados calculados
    for proaq in data:
        obj = ProaqDadosGerais.objects.get(id=proaq['id'])
        proaq['get_status_label'] = obj.get_status_label()
        proaq['get_unidade_daf_label'] = obj.get_unidade_daf_label()
        proaq['get_modalidade_aquisicao_label'] = obj.get_modalidade_aquisicao_label()
        proaq['get_denominacao_nome'] = obj.get_denominacao_nome()
        proaq['get_usuario_nome'] = obj.get_usuario_nome()
    
    return JsonResponse({
        'data': data,
        'total_processos': total_processos,
        'has_next': proaqs_paginados.has_next(),
        'has_previous': proaqs_paginados.has_previous(),
        'current_page': page
    })

@login_required
def salvar_proaq_produto(request, proaq_id):
    if request.method == "POST":
        proaq = get_object_or_404(ProaqDadosGerais, pk=proaq_id)

        # Decodificar os produtos enviados de uma string JSON para uma lista de dicionários
        produtos_selecionados_json = request.POST.get('produtos')
        produtos_selecionados_list = json.loads(produtos_selecionados_json)

        # Obter apenas os IDs dos produtos selecionados para comparação
        produtos_selecionados_ids = [produto["id"] for produto in produtos_selecionados_list]
        
        # Produtos atuais do Proaq
        produtos_banco_ids = list(proaq.proaq_produto.active().values_list('produto_id', flat=True))
        
        # Produtos para salvar
        produtos_salvar = [produto for produto in produtos_selecionados_list if produto["id"] not in produtos_banco_ids]

        # Usuario
        usuario_instance = request.user.usuario_relacionado
        
        # Produtos para adicionar
        for produto_data in produtos_salvar:
            produto_instance = ProaqProdutos(
                usuario_registro=usuario_instance,
                usuario_atualizacao=usuario_instance,
                proaq=proaq,
                produto_id=produto_data["id"]
            )
            produto_instance.save()

        # Produtos para deletar (soft delete)
        produtos_deletar = [produto_id for produto_id in produtos_banco_ids if produto_id not in produtos_selecionados_ids]
        
        for produto_id in produtos_deletar:
            produto_instance = ProaqProdutos.objects.filter(proaq=proaq, produto_id=produto_id, del_status=False).first()
            if produto_instance:
                produto_instance.soft_delete(usuario_instance)

        messages.success(request, f"Produtos atualizados com sucesso!")
        return JsonResponse({"status": "success"})
    else:
        messages.error(request, f"Erro! Produtos não atualizados!")
        return JsonResponse({"status": "error", "message": "Invalid request method"})

def proaq_exportar(request):
    print("Exportar Processos Aquisitivos")
    
    if request.method == 'POST':
        data = json.loads(request.body)
        status_proaq = data.get('status_proaq')
        modalidade_aquisicao = data.get('modalidade_aquisicao')
        unidade_daf = data.get('unidade_daf')
        denominacao = data.get('denominacao')
        
        filters = {}
        filters['del_status'] = False
        if status_proaq:
            filters['status'] = status_proaq
        if modalidade_aquisicao:
            filters['modalidade_aquisicao'] = modalidade_aquisicao
        if unidade_daf:
            filters['unidade_daf'] = unidade_daf
        if denominacao:
            filters['denominacao'] = denominacao

        proaqs = ProaqDadosGerais.objects.filter(**filters)
        data_exportacao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Cria o workbook e adiciona as abas
        wb = Workbook()
        proaq_dados_gerais = wb.active
        proaq_dados_gerais.title = "ProaqDadosGerais"
        proaq_produtos = wb.create_sheet(title="ProaqProdutos")
        proaq_evolucao = wb.create_sheet(title="ProaqEvolucao")
        proaq_tramitacoes = wb.create_sheet(title="ProaqTramitacao")
        
        # Escreve os cabeçalhos
        proaq_dados_gerais.append([
            'ID Proaq', 'Data Registro', 'Responsável Registro', 'Últ. Atualização', 'Responsável Últ. Atualização', 'N Edições',
            'Unidade DAF', 'Modalidade Aquisicao', 'Numero Processo SEI', 'Numero ETP',
            'Status', 'Responsavel Tecnico', 'Denominacao', 'Observacoes Gerais', 'Data Exportação'
        ])
        proaq_produtos.append([
            'ID ProaqProduto', 'ID Proaq', 'Data Registro', 'Responsável Registro', 'Últ. Atualização', 'Responsável Últ. Atualização', 'N Edições',
            'Produto', 'ID Produto', 'Data Exportação'
        ])
        proaq_evolucao.append([
            'ID Evolução', 'ID Proaq', 'Data Registro', 'Responsável Registro', 'Últ. Atualização', 'Responsável Últ. Atualização', 'N Edições',
            'Fase', 'Status', 'Data Inicio', 'Data Fim', 'Comentário'
        ])
        proaq_tramitacoes.append([
            'ID Tramitação', 'ID Proaq', 'Data Registro', 'Responsável Registro', 'Últ. Atualização', 'Responsável Últ. Atualização', 'N Edições',
            'Documento SEI', 'Setor', 'Etapa Processo', 'Data Entrada', 'Previsao Saida', 'Data Saida', 'Observacoes', 'Data Exportação'
        ])
        
        for proaq in proaqs:
            registro_data = proaq.registro_data.replace(tzinfo=None)
            ult_atual_data = proaq.ult_atual_data.replace(tzinfo=None)
            # Escreve os dados em ProaqDadosGerais
            proaq_dados_gerais.append([
                proaq.id,
                registro_data, str(proaq.usuario_registro.primeiro_ultimo_nome()),
                ult_atual_data, str(proaq.usuario_atualizacao.primeiro_ultimo_nome()), proaq.log_n_edicoes,
                proaq.get_unidade_daf_label(), proaq.get_modalidade_aquisicao_label(),
                proaq.numero_processo_sei, proaq.numero_etp, proaq.get_status_label(),
                proaq.get_usuario_nome(), proaq.get_denominacao_nome(), proaq.observacoes_gerais, data_exportacao
            ])
            
            # Escreve os registros relacionados em ProaqProdutos
            for produto in proaq.proaq_produto.filter(del_status=False):
                registro_data = produto.registro_data.replace(tzinfo=None)
                ult_atual_data = produto.ult_atual_data.replace(tzinfo=None)
                proaq_produtos.append([
                    produto.id, produto.proaq.id,
                    registro_data, str(produto.usuario_registro.primeiro_ultimo_nome()),
                    ult_atual_data, str(produto.usuario_atualizacao.primeiro_ultimo_nome()), produto.log_n_edicoes,
                    produto.produto.produto, produto.produto.id, data_exportacao
                ])
            
            # Escreve os registros relacionados em ProaqEvolucao
            for evolucao in proaq.proaq_evolucao.filter(del_status=False):
                registro_data = evolucao.registro_data.replace(tzinfo=None)
                ult_atual_data = evolucao.ult_atual_data.replace(tzinfo=None)
                proaq_evolucao.append([
                    evolucao.id, evolucao.proaq.id,
                    registro_data, str(evolucao.usuario_registro.primeiro_ultimo_nome()),
                    ult_atual_data, str(evolucao.usuario_atualizacao.primeiro_ultimo_nome()), evolucao.log_n_edicoes, 
                    evolucao.fase, evolucao.get_status_display(),
                    evolucao.data_inicio, evolucao.data_fim, evolucao.comentario, data_exportacao
                ])
            
            # Escreve os registros relacionados em ProaqTramitacao
            for tramitacao in proaq.proaq_tramitacao.filter(del_status=False):
                registro_data = proaq.registro_data.replace(tzinfo=None)
                ult_atual_data = proaq.ult_atual_data.replace(tzinfo=None)
                proaq_tramitacoes.append([
                    tramitacao.id, tramitacao.proaq.id, 
                    registro_data, str(tramitacao.usuario_registro.primeiro_ultimo_nome()),
                    ult_atual_data, str(tramitacao.usuario_atualizacao.primeiro_ultimo_nome()), tramitacao.log_n_edicoes,
                    tramitacao.documento_sei, tramitacao.setor,
                    tramitacao.etapa_processo, tramitacao.data_entrada, tramitacao.previsao_saida,
                    tramitacao.data_saida, tramitacao.observacoes, data_exportacao
                ])
        
        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Processo Aquisitivo",
            model='ProaqDadosGerais',
            model_id=0,
            item_id=0,
            item_descricao="Exportação da lista de processos aquisitivos.",
            acao="Exportação",
            observacoes=f"Usuário {request.user.username} exportou dados de Processos Aquisitivos em {current_date_str}."
        )
        log_entry.save()

        # Salva o workbook em um arquivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=proaq.xlsx'
        wb.save(response)
        
        return response

def proaq_tramitacao_delete(request, tramitacao_id=None):
    try:
        tramitacao = ProaqTramitacao.objects.get(id=tramitacao_id)
        tramitacao.soft_delete(request.user.usuario_relacionado)
        messages.error(request, "Tramitação deletada com sucesso!")

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Processo Aquisitivo_Tramitação",
            model='ProaqTramitacao',
            model_id={tramitacao.id},
            item_id=0,
            item_descricao="Deleção de tramitação de processo aquisitivo.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou tramitação (ID: {tramitacao.id}) do processo aquisitivo (ID: {tramitacao.proaq.id}) da unidade daf {tramitacao.proaq.unidade_daf} e denominação genérica {tramitacao.proaq.denominacao.denominacao} em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({"message": "Tramitação deletada com sucesso!"})
    except ProaqTramitacao	.DoesNotExist:
        messages.error(request, "Tramitação não encontrada.")    
    return redirect('proaq')

def proaq_dados_gerais_delete(request, proaq_id=None):
    try:
        proaq = ProaqDadosGerais.objects.get(id=proaq_id)
        proaq.soft_delete(request.user.usuario_relacionado)
        messages.error(request, "Processo Aquisitivo deletado com sucesso.")

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Processo Aquisitivo_Dados Gerais",
            model='ProaqDadosGerais',
            model_id={proaq.id},
            item_id=0,
            item_descricao="Deleção de processo aquisitivo.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o processo aquisitivo (ID: {proaq.id}) da unidade daf {proaq.unidade_daf} e denominação genérica {proaq.denominacao.denominacao} em {current_date_str}."
        )
        log_entry.save()
        
        return JsonResponse({
            "message": "Processo Aquisitivo deletado com sucesso!",
            'redirect_url': '/proaq/',
            })
    except ProaqDadosGerais.DoesNotExist:
        messages.error(request, "Processo Aquisitivo não encontrado.")
        return JsonResponse({"message": "Processo Aquisitivo não encontrado."}) 