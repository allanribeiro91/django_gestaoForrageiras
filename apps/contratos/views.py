from django.shortcuts import render, redirect
from django.db.models import Sum, F, Case, When
from django.db import models
from django.http import QueryDict
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage
from django.contrib import auth, messages
from apps.main.models import CustomLog
from apps.produtos.models import DenominacoesGenericas, ProdutosFarmaceuticos
from apps.fornecedores.models import Fornecedores
from apps.contratos.models import ContratosArps, ContratosArpsItens, Contratos
from apps.contratos.forms import ContratosArpsForm, ContratosArpsItensForm, ContratosForm
from setup.choices import UNIDADE_DAF, MODALIDADE_AQUISICAO, STATUS_ARP, YES_NO, TIPO_COTA
from django.http import JsonResponse, HttpResponse
from datetime import datetime
from django.utils import timezone
import pytz
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import json

#timezone
tz = pytz.timezone("America/Sao_Paulo")


#CONTRATOS
def contratos(request):

    lista_modalidades = [('', '')] + MODALIDADE_AQUISICAO
    lista_unidadesdaf = [item for item in UNIDADE_DAF if item[0] not in ['cofisc', 'gabinete']]
    
    tabContratos = Contratos.objects.filter(del_status=False)
    
    conteudo = {
        'lista_unidadesdaf': lista_unidadesdaf,
        'lista_modalidades': lista_modalidades,
        'tabContratos': tabContratos,
    }
    return render(request, 'contratos/contratos.html', conteudo)

def contrato_ficha(request, id_contrato=None):
    if id_contrato:
        contrato = Contratos.objects.get(id=id_contrato)
    else:
        contrato = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if contrato:
            contrato_form = ContratosForm(request.POST, instance=contrato)
            novo_contrato = False
        else:
            contrato_form = ContratosForm(request.POST)
            novo_contrato = True
        
        #Verificar se houve alteração no formulário
        if not contrato_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })

        #Passar o objeto Denominação Genérica
        denominacao_id = request.POST.get('denominacao')
        print('Denominação ID: ', denominacao_id)
        denominacao_instance = DenominacoesGenericas.objects.get(id=denominacao_id)
    
        #Passar o objeto Fornecedor
        fornecedor_id = request.POST.get('fornecedor')
        fornecedor_instance =  Fornecedores.objects.get(id=fornecedor_id)

        #Passar a ARP
        modalidade = request.POST.get('modalidade_aquisicao')
        if modalidade == 'pregao_comarp':
            arp_id = request.POST.get('arp')
            arp_instance = ContratosArps.objects.get(id=arp_id)
        else:
            arp_instance = None

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #lei de licitacao
        lei_licitacao = request.POST.get('ct_lei_licitacao_valor')

        #Atualizar os valores no mutable_post
        modificacoes_post['denominacao'] = denominacao_instance
        modificacoes_post['fornecedor'] = fornecedor_instance
        modificacoes_post['arp'] = arp_instance
        modificacoes_post['lei_licitacao'] = lei_licitacao

        #Criar o formulário com os dados atualizados
        contrato_form = ContratosForm(modificacoes_post, instance=contrato_form.instance)
        
        #salvar
        if contrato_form.is_valid():
            #Salvar o produto
            contrato = contrato_form.save(commit=False)
            contrato.save(current_user=request.user.usuario_relacionado)
            
            #logs
            log_id = contrato.id
            log_registro_usuario = contrato.usuario_registro.dp_nome_completo
            log_registro_data = contrato.registro_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_atualizacao_usuario = contrato.usuario_atualizacao.dp_nome_completo
            log_atualizacao_data = contrato.ult_atual_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_edicoes = contrato.log_n_edicoes

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_Contratos",
                model='Contratos',
                model_id=contrato.id,
                item_id=0,
                item_descricao="Salvar edição de Contrato.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o Contrato (ID {contrato.id}, Número do Contrato: {contrato.numero_contrato}, Denominação: {contrato.denominacao.denominacao}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'log_id': log_id,
                    'log_registro_usuario': log_registro_usuario,
                    'log_registro_data': log_registro_data,
                    'log_atualizacao_usuario': log_atualizacao_usuario,
                    'log_atualizacao_data': log_atualizacao_data,
                    'log_edicoes': log_edicoes,
                    'novo': novo_contrato,
                    'redirect_url': reverse('contrato_ficha', args=[contrato.id]),
                })
        else:
            print("Erro formulário Contrato")
            print(contrato_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })

    if contrato:
        form = ContratosForm(instance=contrato)
    else:
        form = ContratosForm()

    return render(request, 'contratos/contrato_ficha.html', {
        'form': form,
        'contrato': contrato,
    })

def contrato_delete(request, id_contrato=None):   
    try:
        contrato = Contratos.objects.get(id=id_contrato)
        contrato.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_Contratos",
            model='Contratos',
            model_id=contrato.id,
            item_id=0,
            item_descricao="Deleção de Contrato.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o Contrato (ID {contrato.id}, Número: {contrato.numero_contrato}, Denominação: {contrato.denominacao.denominacao}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Contrato deletado com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "Contrato não encontrado."
            })

def buscar_arps(request, unidade_daf=None):
    arps = ContratosArps.objects.filter(del_status=False, unidade_daf=unidade_daf, status='vigente').order_by('numero_arp')
    arps_list = list(arps.values('id', 'numero_arp', 'denominacao', 'fornecedor'))  # Convertendo para uma lista de dicionários
    return JsonResponse({'arps': arps_list})

def buscar_arps_itens(request, id_arp=None):
    arps_itens = ContratosArpsItens.objects.filter(del_status=False, status='vigente', arp_id=id_arp).order_by('numero_item')
    arps_itens_list = []

    for item in arps_itens:
        arps_item_data = {
            'id': item.id,
            'numero_item': item.numero_item,
            'tipo_cota': item.tipo_cota,
            'produto': item.produto.produto,
            'qtd_registrada': item.qtd_registrada,
            'qtd_saldo': item.qtd_saldo()
        }
        arps_itens_list.append(arps_item_data)

    return JsonResponse({'arps_itens': arps_itens_list})


def buscar_contrato(request, id_contrato=None):
    contrato = Contratos.objects.get(id=id_contrato)
    ct_denominacao_id = contrato.denominacao.id
    ct_denominacao = str(contrato.denominacao)
    ct_fornecedor_id = contrato.fornecedor.id
    ct_fornecedor = str(contrato.fornecedor)
    contrato_dados = {
            'denominacao_id': ct_denominacao_id,
            'denominacao_texto': ct_denominacao,
            'fornecedor_id': ct_fornecedor_id,
            'fornecedor_texto': ct_fornecedor,
        }
    print(contrato_dados)
    return JsonResponse({'contrato': contrato_dados})



#ARPs
def arps(request):
    tab_arps = ContratosArps.objects.all().filter(del_status=False).order_by('-data_publicacao')
    denominacoes = DenominacoesGenericas.objects.values_list('id', 'denominacao')
    fornecedores = Fornecedores.objects.values_list('nome_fantasia', flat=True).distinct().order_by('nome_fantasia')
    unidades_daf = [item for item in UNIDADE_DAF if item[0] not in ['cofisc', 'gabinete']]
    conteudo = {
        'lista_status': STATUS_ARP,
        'lista_unidadesdaf': unidades_daf,
        'lista_denominacoes': denominacoes,
        'lista_fornecedores': fornecedores,
        'tab_arps': tab_arps,
    }
    return render(request, 'contratos/arps.html', conteudo)

def arp_ficha(request, arp_id=None):
    if arp_id:
        try:
            arp = ContratosArps.objects.get(id=arp_id)
        except ContratosArps.DoesNotExist:
            messages.error(request, "ARP não encontrada.")
            return redirect('arps')
    else:
        arp = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if arp:
            arp_form = ContratosArpsForm(request.POST, instance=arp)
            nova_arp = False
        else:
            arp_form = ContratosArpsForm(request.POST)
            nova_arp = True

        #Verificar se houve alteração no formulário
        if not arp_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })

        #Passar o objeto Denominação Genérica
        denominacao_id = request.POST.get('denominacao')
        if denominacao_id == None:
            denominacao_id = request.POST.get('arp_denominacao_hidden')
        denominacao_instance = DenominacoesGenericas.objects.get(id=denominacao_id)
    
        #Passar o objeto Fornecedor
        fornecedor_id = request.POST.get('fornecedor')
        fornecedor_instance =  Fornecedores.objects.get(id=fornecedor_id)
        
        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Atualizar os valores no mutable_post
        modificacoes_post['denominacao'] = denominacao_instance
        modificacoes_post['fornecedor'] = fornecedor_instance

        #Criar o formulário com os dados atualizados
        arp_form = ContratosArpsForm(modificacoes_post, instance=arp_form.instance)

        #salvar
        if arp_form.is_valid():
            #Salvar o produto
            arp = arp_form.save(commit=False)
            arp.save(current_user=request.user.usuario_relacionado)
            
            #logs
            log_id = arp.id
            log_registro_usuario = arp.usuario_registro.dp_nome_completo
            log_registro_data = arp.registro_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_atualizacao_usuario = arp.usuario_atualizacao.dp_nome_completo
            log_atualizacao_data = arp.ult_atual_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_edicoes = arp.log_n_edicoes

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_ARPs",
                model='ContratosARPs',
                model_id=arp.id,
                item_id=0,
                item_descricao="Salvar edição de ARP.",
                acao="Salvar",
                #observacoes=f"Usuário {request.user.username} salvou a ARP (ID {arp.id}, Número: {arp.topico}, Denominação: {arp.denominacao.denominacao}, Fornecedor: {arp.fornecedor.nome_fantasia}) em {current_date_str}."
                observacoes=f"Usuário {request.user.username} salvou a ARP (ID {arp.id}, Número: {arp.numero_arp}, Denominação: {arp.denominacao.denominacao}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'log_id': log_id,
                    'log_registro_usuario': log_registro_usuario,
                    'log_registro_data': log_registro_data,
                    'log_atualizacao_usuario': log_atualizacao_usuario,
                    'log_atualizacao_data': log_atualizacao_data,
                    'log_edicoes': log_edicoes,
                    'novo': nova_arp,
                    'redirect_url': reverse('arp_ficha', args=[arp.id]),
                })
        else:
            print("Erro formulário ARP")
            print(arp_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })
            
    #Form ARP
    valor_total_arp = 0
    tab_itens_arp = None
    qtd_registrada_total_arp = 0
    qtd_saldo_total_arp = 0
    qtd_saldo_total_arp_percentual = 0
    if arp:
        form = ContratosArpsForm(instance=arp)
        
        #Itens da ARP
        tab_itens_arp = ContratosArpsItens.objects.filter(del_status=False, arp_id=arp.id).order_by('numero_item')
        
        #Calculando o valor total para cada item
        tab_itens_arp = tab_itens_arp.annotate(
            valor_total=Case(
                When(valor_unit_reequilibrio_bool=True, then=F('valor_unit_reequilibrio') * F('qtd_registrada')),
                default=F('valor_unit_homologado') * F('qtd_registrada'),
                output_field=models.FloatField()
                )
            )
        
        #valores
        valor_total_arp = tab_itens_arp.aggregate(total=Sum('valor_total'))['total']
        if valor_total_arp == None:
            valor_total_arp = 0
        
        qtd_registrada_total_arp = tab_itens_arp.aggregate(total=Sum('qtd_registrada'))['total']
        if qtd_registrada_total_arp == None:
            qtd_registrada_total_arp = 0
        
        qtd_saldo_total_arp = 0
        for item in tab_itens_arp:
            qtd_saldo_total_arp += item.qtd_saldo()

        if qtd_saldo_total_arp > 0:
            qtd_saldo_total_arp_percentual = qtd_saldo_total_arp / qtd_registrada_total_arp
    else:
        form = ContratosArpsForm()

    #Form Item da ARP
    form_item = ContratosArpsItensForm()

    return render(request, 'contratos/arp_ficha.html', {
        'YES_NO': YES_NO,
        'TIPO_COTA': TIPO_COTA,
        'form': form,
        'form_item': form_item,
        'arp': arp,
        'tab_itens_arp': tab_itens_arp,
        'valor_total_arp': valor_total_arp,
        'qtd_registrada_total_arp': qtd_registrada_total_arp,
        'qtd_saldo_total_arp': qtd_saldo_total_arp,
        'qtd_saldo_total_arp_percentual': qtd_saldo_total_arp_percentual,
    })

def arp_delete(request, arp_id=None):   
    try:
        arp = ContratosArps.objects.get(id=arp_id)
        arp.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_ARPs",
            model='ContratosARPs',
            model_id=arp.id,
            item_id=0,
            item_descricao="Deleção de ARP.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou a ARP (ID {arp.id}, Número: {arp.numero_arp}, Denominação: {arp.denominacao.denominacao}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "ARP deletada com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "ARP não encontrada."
            })

def arp_buscar_produtos(request, denominacao=None):
    produtos = ProdutosFarmaceuticos.get_produtos_por_denominacao(denominacao)
    return JsonResponse(produtos, safe=False)

def arp_filtrar(request):
    status_arp = request.GET.get('status_arp', None)
    unidade_daf = request.GET.get('unidade_daf', None)
    denominacao = request.GET.get('denominacao', None)
    fornecedor = request.GET.get('fornecedor', None)

    filters = {}
    filters['del_status'] = False
    if status_arp:
        filters['status'] = status_arp
    if unidade_daf:
        filters['unidade_daf'] = unidade_daf
    if denominacao:
        filters['denominacao_id'] = denominacao
    if fornecedor:
        filters['fornecedor__nome_fantasia__icontains'] = fornecedor
    
    tab_arps = ContratosArps.objects.filter(**filters).order_by('-data_publicacao')
    total_arps = tab_arps.count()

    page = int(request.GET.get('page', 1))
    paginator = Paginator(tab_arps, 100)  # Mostra 100 faqs por página
    try:
        arp_paginados = paginator.page(page)
    except EmptyPage:
        arp_paginados = paginator.page(paginator.num_pages)

    data = []
    for arp in arp_paginados.object_list:
        valor_total = arp.valor_total_arp()
        arp_data = {
            'id': arp.id,
            'status': arp.status,
            'unidade_daf': arp.unidade_daf,
            'numero_processo_sei': arp.numero_processo_sei,
            'numero_documento_sei': arp.numero_documento_sei,
            'data_publicacao': arp.data_publicacao.strftime('%d/%m/%Y') if arp.data_publicacao else '',
            'data_vigencia': arp.data_vigencia.strftime('%d/%m/%Y') if arp.data_publicacao else '',
            'prazo_vigencia': arp.prazo_vigencia if arp.data_publicacao else '',
            'denominacao': arp.denominacao.denominacao,
            'fornecedor': arp.fornecedor.nome_fantasia,
            'valor_total_arp': valor_total,
        }
        data.append(arp_data)

    return JsonResponse({
        'data': data,
        'total_arps': total_arps,
        'has_next': arp_paginados.has_next(),
        'has_previous': arp_paginados.has_previous(),
        'current_page': page
    })

def arp_exportar(request):
    print("Exportar ARPs")
    
    if request.method == 'POST':
        data = json.loads(request.body)
        status_arp = data.get('status_arp')
        unidade_daf = data.get('unidade_daf')
        denominacao = data.get('denominacao')
        fornecedor = data.get('fornecedor')

        filters = {}
        filters['del_status'] = False
        if status_arp:
            filters['status'] = status_arp
        if unidade_daf:
            filters['unidade_daf'] = unidade_daf
        if denominacao:
            filters['denominacao_id'] = denominacao
        if fornecedor:
            filters['fornecedor__nome_fantasia__icontains'] = fornecedor
        
        arps = ContratosArps.objects.filter(**filters)
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Criar um workbook e adicionar uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "arps"

        headers = [
        'ID', 'Usuário Registro', 'Usuário Atualização', 'Data de Registro', 'Data da Última Atualização', 'N Edições', 
        'Unidade DAF', 'Processo SEI', 'Documento SEI', 'ARP', 'Status', 'Data Publicacao',
        'Denominacao', 'Fornecedor',
        'Observações Gerais', 'Data Exportação'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws['{}1'.format(col_letter)] = header
            ws.column_dimensions[col_letter].width = 20

        # Adicionar dados da tabela
        for row_num, arp in enumerate(arps, 2):
            registro_data = arp.registro_data.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M:%S')
            ult_atual_data = arp.ult_atual_data.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M:%S')
            data_publicacao = arp.data_publicacao
            if data_publicacao:
                data_publicacao.strftime('%d/%m/%Y')
            denominacao = str(arp.denominacao)
            fornecedor = str(arp.fornecedor)

            ws.cell(row=row_num, column=1, value=arp.id)
            ws.cell(row=row_num, column=2, value=str(arp.usuario_registro.primeiro_ultimo_nome()))
            ws.cell(row=row_num, column=3, value=str(arp.usuario_atualizacao.primeiro_ultimo_nome()))
            ws.cell(row=row_num, column=4, value=registro_data)
            ws.cell(row=row_num, column=5, value=ult_atual_data)
            ws.cell(row=row_num, column=6, value=arp.log_n_edicoes)
            ws.cell(row=row_num, column=7, value=arp.unidade_daf)
            ws.cell(row=row_num, column=8, value=arp.numero_processo_sei)
            ws.cell(row=row_num, column=9, value=arp.numero_documento_sei)
            ws.cell(row=row_num, column=10, value=arp.numero_arp)
            ws.cell(row=row_num, column=11, value=arp.status)
            ws.cell(row=row_num, column=12, value=data_publicacao)
            ws.cell(row=row_num, column=13, value=denominacao)
            ws.cell(row=row_num, column=14, value=fornecedor)
            ws.cell(row=row_num, column=15, value=arp.observacoes_gerais)
            ws.cell(row=row_num, column=16, value=current_date_str)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Reposition to the start of the stream

        # Registrar a ação no CustomLog
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_ARPs",
            model='ContratosARPs',
            model_id=0,
            item_id=0,
            item_descricao="Exportação da lista de ARPs",
            acao="Exportação",
            observacoes=f"Usuário {request.user.username} exportou lista de ARPs em {current_date_str}."
        )
        log_entry.save()

        # Configurar a resposta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportar_fornecedores.xlsx"'
        response.write(output.getvalue())
        return response

def arp_buscar_dados_sei(request, id_arp=None):
    arp = ContratosArps.objects.filter(id=id_arp)
    arp_list = list(arp.values('id', 'numero_processo_sei', 'lei_licitacao'))
    print(arp_list)
    return JsonResponse({'arp': arp_list})


#ITENS DAS ARPS
def arp_item_ficha(request, arp_item_id=None):
    if arp_item_id:
        try:
            item_arp = ContratosArpsItens.objects.get(id=arp_item_id)
        except ContratosArpsItens.DoesNotExist:
            messages.error(request, "Item da ARP não encontrada.")
            return redirect('arps')
    else:
        item_arp = None
    
    #salvar
    if request.method == 'POST':
        #Carregar formulário
        if item_arp:
            item_arp_form = ContratosArpsItensForm(request.POST, instance=item_arp)
            novo_item_arp = False
        else:
            item_arp_form = ContratosArpsItensForm(request.POST)
            novo_item_arp = True

        #Verificar se houve alteração no formulário
        if not item_arp_form.has_changed():
            return JsonResponse({
                    'retorno': 'Não houve mudanças'
                })

        #Fazer uma cópia mutável do request.POST
        modificacoes_post = QueryDict(request.POST.urlencode(), mutable=True)

        #Passar o objeto ARP
        arp_id = request.POST.get('id_arp')
        arp_instance = ContratosArps.objects.get(id=arp_id)

        #Passar o objeto Produto Farmacêutico
        produto_id = request.POST.get('produto')
        produto_instance = ProdutosFarmaceuticos.objects.get(id=produto_id)

        #valor unitario homologado
        valor_homologado_str  = request.POST.get('valor_unit_homologado')
        valor_homologado_str = valor_homologado_str.replace('R$', '').replace('.', '')
        valor_homologado_str = valor_homologado_str.replace(',', '.')
        valor_unit = float(valor_homologado_str)

        #valor unitario reequilibrio
        valor_reequilibrio_str  = request.POST.get('valor_unit_reequilibrio')
        print('valor = ', valor_reequilibrio_str)
        if valor_reequilibrio_str != "":
            print('teste')
            valor_reequilibrio_str = valor_reequilibrio_str.replace('R$', '').replace('.', '')
            valor_reequilibrio_str = valor_reequilibrio_str.replace(',', '.')
            valor_reequilibrio = float(valor_reequilibrio_str)
            modificacoes_post['valor_unit_reequilibrio'] = valor_reequilibrio

        #qtd registrada
        qtd_registrada_str = request.POST.get('qtd_registrada')
        qtd_registrada_str = qtd_registrada_str.replace('.', '')
        qtd_registrada_int = int(qtd_registrada_str)

        #Atualizar os valores no mutable_post
        modificacoes_post['arp'] = arp_instance
        modificacoes_post['produto'] = produto_instance
        modificacoes_post['valor_unit_homologado'] = valor_unit
        modificacoes_post['qtd_registrada'] = qtd_registrada_int

        #Criar o formulário com os dados atualizados
        item_arp_form = ContratosArpsItensForm(modificacoes_post, instance=item_arp_form.instance)

        #salvar
        if item_arp_form.is_valid():
            #Salvar o produto
            item_arp = item_arp_form.save(commit=False)
            item_arp.save(current_user=request.user.usuario_relacionado)
            
            #logs
            log_id = item_arp.id
            log_atualizacao_usuario = item_arp.usuario_atualizacao.dp_nome_completo
            log_atualizacao_data = item_arp.ult_atual_data.astimezone(tz).strftime('%d/%m/%Y %H:%M:%S')
            log_edicoes = item_arp.log_n_edicoes

            # Registrar a ação no CustomLog
            current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_entry = CustomLog(
                usuario=request.user.usuario_relacionado,
                modulo="Contratos_ARPs_Itens",
                model='ContratosARPsItens',
                model_id=item_arp.id,
                item_id=0,
                item_descricao="Salvar edição de Item da ARP.",
                acao="Salvar",
                observacoes=f"Usuário {request.user.username} salvou o Item da ARP (ID {item_arp.id}, Nº Item: {item_arp.numero_item}, Nº ARP: {item_arp.arp.numero_arp}, Produto: {item_arp.produto.produto}) em {current_date_str}."
            )
            log_entry.save()

            #Retornar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'retorno': 'Salvo',
                    'log_id': log_id,
                    'log_atualizacao_usuario': log_atualizacao_usuario,
                    'log_atualizacao_data': log_atualizacao_data,
                    'log_edicoes': log_edicoes,
                    'novo': novo_item_arp,
                    'redirect_url': reverse('arp_ficha', args=[item_arp.arp_id]),
                })
        else:
            print("Erro formulário Item da ARP")
            print(item_arp_form.errors)
            return JsonResponse({
                    'retorno': 'Erro ao salvar'
                })
            
    #Form ARP
    if item_arp:
        form_item = ContratosArpsItensForm(instance=item_arp)
    else:
        form_item = ContratosArpsItensForm()

    return render(request, 'contratos/arp_ficha.html', {
        'YES_NO': YES_NO,
        'TIPO_COTA': TIPO_COTA,
        'form_item': form_item,
        'item_arp': item_arp,
    })

def arp_item_formulario(request, arp_item_id=None):
    try:
        item = ContratosArpsItens.objects.get(id=arp_item_id)
        print('qtd: ',float(item.qtd_registrada))
        produto_id = item.produto_id
        produto_nome = item.produto.produto
        data = {
            'id': item.id,
            'log_data_registro': item.registro_data.strftime('%d/%m/%Y %H:%M:%S') if item.registro_data else '',
            'log_responsavel_registro': str(item.usuario_atualizacao.dp_nome_completo),
            'lot_ult_atualizacao': item.ult_atual_data.strftime('%d/%m/%Y %H:%M:%S') if item.ult_atual_data else '',
            'log_responsavel_atualizacao': str(item.usuario_atualizacao.dp_nome_completo),
            'log_edicoes': item.log_n_edicoes,
            'numero_item': item.numero_item,
            'tipo_cota': item.tipo_cota,
            'empate_ficto': item.empate_ficto,
            'produto': produto_id,
            'valor_unit_homologado': item.valor_unit_homologado,
            'valor_unit_reequilibrio_bool': item.valor_unit_reequilibrio_bool,
            'valor_unit_reequilibrio': item.valor_unit_reequilibrio,
            'qtd_registrada': item.qtd_registrada,
            'observacoes': item.observacoes_gerais if item.observacoes_gerais else '',
        }
        return JsonResponse(data)
    except ContratosArpsItens.DoesNotExist:
        return JsonResponse({'error': 'Item da ARP não encontrada'}, status=404)

def arp_item_delete(request, arp_item_id=None):   
    try:
        item_arp = ContratosArpsItens.objects.get(id=arp_item_id)
        item_arp.soft_delete(request.user.usuario_relacionado)

        # Registrar a ação no CustomLog
        current_date_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        log_entry = CustomLog(
            usuario=request.user.usuario_relacionado,
            modulo="Contratos_ARPs_Itens",
            model='ContratosARPsItens',
            model_id=item_arp.id,
            item_id=0,
            item_descricao="Deleção de Itens da ARP.",
            acao="Deletar",
            observacoes=f"Usuário {request.user.username} deletou o Item da ARP (ID {item_arp.id}, Nº Item: {item_arp.numero_item}, Protudo: {item_arp.produto.produto}) em {current_date_str}."
        )
        log_entry.save()

        return JsonResponse({
            "message": "Item da ARP deletada com sucesso!"
            })
    except ContratosArps.DoesNotExist:
        return JsonResponse({
            "message": "Item da ARP não encontrada."
            })

